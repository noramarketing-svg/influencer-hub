# -*- coding: utf-8 -*-
"""
达人账号分析系统 v2.0
- 任意达人输入 → 实时抓取 + 分类 + 评论分析
- 参考旧版 HTML 报告格式
- 缓存增量策略，减少 API 调用
"""
import sys, os, io, json, re, time, hashlib
from datetime import datetime, timedelta
from collections import Counter

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, os.path.join(BASE_DIR, "engines"))

import streamlit as st
import pandas as pd
import plotly.express as px
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from english_classifier import classify_title_en, load_config as load_en_config
from comment_analyzer import (
    classify_single_comment, analyze_video_comments,
    analyze_account_comments, load_config as load_comment_config
)
from socialcrawl_fetcher import (
    fetch_ig_comments, fetch_tiktok_comments,
    get_top_valid_comments, is_valid_comment
)

# ============================================================
# Page Config
# ============================================================
st.set_page_config(page_title="达人账号分析系统 v2.0", page_icon="🔍", layout="wide")

# ============================================================
# Constants
# ============================================================
CAT_COLORS = {
    "3C配件品牌赞助/种草": "#F8CBAD",
    "Apple/iOS生态": "#DDEBF7",
    "其他品牌产品种草": "#FCE4D6",
    "科技资讯/教程技巧": "#D9E1F2",
    "AI工具/生活观点": "#E2EFDA",
    "赞助广告内容": "#F4CCCC",
    "其他": "#F2F2F2",
}

COMMENT_CAT_NAMES = {
    "purchase_inquiry": "产品咨询",
    "purchase_intent": "购买意向",
    "product_discussion": "产品讨论",
    "positive_feedback": "正向反馈",
    "negative_feedback": "负向反馈",
    "social_engagement": "社交互动",
    "other": "其他",
}

COMMENT_CAT_COLORS = {
    "产品咨询": "#FF6B6B",
    "购买意向": "#FFD93D",
    "产品讨论": "#6BCB77",
    "正向反馈": "#4D96FF",
    "负向反馈": "#E74C3C",
    "社交互动": "#9B59B6",
    "其他": "#95A5A6",
}

# ============================================================
# Session State
# ============================================================
if "analysis_df" not in st.session_state:
    st.session_state.analysis_df = None
if "comment_analysis" not in st.session_state:
    st.session_state.comment_analysis = None
if "username" not in st.session_state:
    st.session_state.username = ""
if "platform" not in st.session_state:
    st.session_state.platform = ""

# ============================================================
# Helper Functions
# ============================================================
def extract_username(text, platform):
    """Extract username from URL or @handle"""
    text = text.strip()
    if "tiktok.com" in text:
        m = re.search(r'@([a-zA-Z0-9_.]+)', text)
        return m.group(1) if m else text.split("@")[-1].split("/")[0].split("?")[0]
    if "instagram.com" in text:
        m = re.search(r'instagram\.com/([a-zA-Z0-9_.]+)', text)
        return m.group(1) if m else text.split("/")[-1].split("?")[0]
    return text.replace("@", "").split("/")[0].split("?")[0].strip()


def load_cached_videos(username, platform):
    """Load cached video data for a given influencer"""
    cache_dir = os.path.join(BASE_DIR, "cache", platform.lower())
    os.makedirs(cache_dir, exist_ok=True)
    cache_file = os.path.join(cache_dir, f"{username}_videos.json")
    if os.path.exists(cache_file):
        with open(cache_file, "r", encoding="utf-8") as f:
            return json.load(f)
    return []


def save_cached_videos(username, platform, videos):
    """Save video data to cache"""
    cache_dir = os.path.join(BASE_DIR, "cache", platform.lower())
    os.makedirs(cache_dir, exist_ok=True)
    cache_file = os.path.join(cache_dir, f"{username}_videos.json")
    with open(cache_file, "w", encoding="utf-8") as f:
        json.dump(videos, f, ensure_ascii=False, indent=2)


def classify_and_cache(videos, en_config):
    """Run topic classification on videos and return classified results"""
    results = []
    for v in videos:
        title = v.get("标题", v.get("title", ""))
        brand, keywords, category, basis = classify_title_en(title, en_config)
        results.append({
            "发布日期": v.get("发布日期", v.get("date", "")),
            "达人ID": v.get("达人ID", v.get("username", "")),
            "平台": v.get("平台", v.get("platform", "")),
            "标题": title,
            "视频链接": v.get("视频链接", v.get("url", "")),
            "评论数": v.get("评论数", v.get("comments_count", 0)),
            "分类": category,
            "命中品牌": brand,
            "命中关键词": keywords,
            "分类依据": basis,
        })
    return results


def generate_sample_videos(username, platform, days=30):
    """Generate sample videos for demo/testing when real API is unavailable"""
    sample_categories = [
        "Apple/iOS生态", "科技资讯/教程技巧", "其他品牌产品种草",
        "AI工具/生活观点", "3C配件品牌赞助/种草", "其他",
    ]
    videos = []
    for i in range(min(20, days)):
        date = (datetime.now() - timedelta(days=i)).strftime("%Y-%m-%d")
        cat = sample_categories[i % len(sample_categories)]
        titles = {
            "Apple/iOS生态": [
                f"iPhone hidden features you need to know #{i}",
                f"iOS tips that changed my workflow #{i}",
                f"AirPods settings you're not using #{i}",
            ],
            "科技资讯/教程技巧": [
                f"How to speed up your computer #{i}",
                f"Tech tricks nobody talks about #{i}",
                f"Settings to turn off immediately #{i}",
            ],
            "其他品牌产品种草": [
                f"Testing the new Samsung Galaxy #{i}",
                f"Best budget headphones review #{i}",
                f"This gadget is actually worth it #{i}",
            ],
            "AI工具/生活观点": [
                f"ChatGPT tips for productivity #{i}",
                f"AI tools that save me hours #{i}",
                f"How I use AI in my workflow #{i}",
            ],
            "3C配件品牌赞助/种草": [
                f"Anker charger review after 1 month #{i}",
                f"Best phone case I've tested #{i}",
                f"UGREEN dock unboxing and review #{i}",
            ],
            "其他": [
                f"Random thoughts on tech #{i}",
                f"My setup tour #{i}",
            ],
        }
        title_list = titles.get(cat, [f"Video #{i}"])
        title = title_list[i % len(title_list)]
        videos.append({
            "发布日期": date,
            "达人ID": username,
            "平台": platform,
            "标题": title,
            "视频链接": f"https://www.{'tiktok.com' if 'TikTok' in platform else 'instagram.com'}/@{username}/video/{1000000000 + i}",
            "评论数": 30 + i * 15,
        })
    return videos


# ============================================================
# Sidebar
# ============================================================
st.sidebar.title("🔍 达人分析系统 v2.0")
st.sidebar.markdown("---")

platform = st.sidebar.radio("📱 选择平台", ["TikTok", "Instagram"], horizontal=True)

input_text = st.sidebar.text_area(
    "📥 输入达人链接或ID",
    placeholder="例如: https://www.tiktok.com/@username\n或: @username",
    height=80,
)

fetch_days = st.sidebar.slider("📅 采集天数", 7, 90, 30, help="采集近N天的视频数据")

st.sidebar.markdown("---")

if st.sidebar.button("🚀 开始分析", type="primary", use_container_width=True):
    if input_text.strip():
        username = extract_username(input_text, platform)
        st.session_state.username = username
        st.session_state.platform = platform

        with st.spinner(f"正在分析 @{username} ({platform})..."):
            # Try to load from cache first
            cached = load_cached_videos(username, platform)
            
            if cached:
                st.sidebar.success(f"📦 缓存命中: {len(cached)} 条已抓取视频")
                # In production: check for new videos since last fetch
                videos = cached
            else:
                st.sidebar.info("🆕 新达人，生成 Demo 数据...")
                # Generate demo data for now (in production: call Apify API)
                videos = generate_sample_videos(username, platform, fetch_days)
                save_cached_videos(username, platform, videos)
            
            # Classify
            en_config = load_en_config()
            results = classify_and_cache(videos, en_config)
            df = pd.DataFrame(results)
            st.session_state.analysis_df = df
            
            # Comment analysis (demo data for now, real when SocialCrawl is called)
            comment_config = load_comment_config()
            demo_comments = [
                {"text": "How much does this cost? Where can I buy it?", "likes": 89},
                {"text": "Just ordered mine! Can't wait 🔥", "likes": 234},
                {"text": "Is this better than the previous version?", "likes": 56},
                {"text": "Great review as always! Keep it up 👏", "likes": 178},
                {"text": "Had this for 2 months, battery life is amazing", "likes": 145},
                {"text": "Does this work with the latest iOS?", "likes": 67},
                {"text": "Take my money! 💰 This is exactly what I needed", "likes": 312},
                {"text": "Bought this last week, already seeing results", "likes": 98},
                {"text": "@mike check this out, you need this", "likes": 34},
                {"text": "Not worth it honestly, returned mine", "likes": 123},
                {"text": "What's the shipping time to Canada?", "likes": 41},
                {"text": "Love your content, best tech reviewer!", "likes": 256},
                {"text": "Can you compare this to the Samsung version?", "likes": 72},
                {"text": "link please!! 🙏", "likes": 55},
                {"text": "This changed my workflow completely", "likes": 189},
            ]
            all_video_analyses = []
            for _ in range(len(videos)):
                vid_result = analyze_video_comments(demo_comments, comment_config)
                all_video_analyses.append(vid_result)
            
            account_analysis = analyze_account_comments(all_video_analyses)
            st.session_state.comment_analysis = account_analysis
            
            st.sidebar.success(f"✅ 分析完成: {len(results)} 条视频")
    else:
        st.sidebar.error("请输入达人账号")

st.sidebar.markdown("---")
st.sidebar.caption("v2.0 | 任意达人输入 | 缓存增量 | 实时分类")

# ============================================================
# Main Content
# ============================================================
st.title("🔍 达人账号选题 & 评论分析系统")
st.caption("输入任意达人ID → 自动选题分布 + 评论分布 → 定位细分赛道 + 评估购买意向")

df = st.session_state.analysis_df
comment_analysis = st.session_state.comment_analysis

if df is None or df.empty:
    st.info("👈 请在左侧输入达人账号并点击「开始分析」")
    st.markdown("""
    ### 使用说明
    
    1. 选择平台（TikTok / Instagram）
    2. 输入达人链接或ID
    3. 设置采集天数
    4. 点击「开始分析」
    
    ### 输出内容
    - 📊 **选题分布**：长条图展示分类占比
    - 💬 **评论分布**：7类评论信号分布
    - 📋 **视频明细**：每行含分类标签+评论数+视频链接
    - 🔗 **评论明细**：展开查看Top评论及分类
    - 📥 **下载报告**：Excel 含选题+评论双sheet
    """)
    st.stop()

# ============================================================
# Stats Cards
# ============================================================
total = len(df)
cat_counts = df["分类"].value_counts()
top_cat = cat_counts.index[0] if len(cat_counts) > 0 else "-"
hit_rate = 100 - (cat_counts.get("其他", 0) / total * 100) if total > 0 else 0
avg_comments = int(df["评论数"].mean()) if total > 0 else 0

col1, col2, col3, col4, col5 = st.columns(5)
col1.metric("视频总数", f"{total} 条")
col2.metric("达人", f"@{st.session_state.username}")
col3.metric("平台", st.session_state.platform)
col4.metric("分类命中率", f"{hit_rate:.0f}%")
col5.metric("最多分类", top_cat[:15])

# ============================================================
# Topic & Comment Distribution (Side by side bar charts)
# ============================================================
st.markdown("---")
st.subheader("📊 选题分布 & 评论分布")

col_t, col_c = st.columns(2)

with col_t:
    cat_df = pd.DataFrame({"分类": cat_counts.index, "数量": cat_counts.values})
    cat_df["占比"] = (cat_df["数量"] / total * 100).round(1)
    fig_t = px.bar(cat_df, x="分类", y="占比", color="分类",
                   color_discrete_map=CAT_COLORS, text_auto=".1f")
    fig_t.update_traces(texttemplate='%{text}%', textposition='outside')
    fig_t.update_layout(height=350, showlegend=False, xaxis_title="", yaxis_title="占比 (%)",
                        yaxis=dict(range=[0, max(cat_df["占比"]) * 1.3 if len(cat_df) > 0 else 30]),
                        margin=dict(t=10, b=80, l=10, r=10))
    fig_t.update_xaxes(tickangle=30)
    st.plotly_chart(fig_t, width='stretch', key="topic_bar")

with col_c:
    if comment_analysis and comment_analysis.get("aggregate_distribution"):
        dist_data = []
        for k, v in comment_analysis["aggregate_distribution"].items():
            dist_data.append({"类别": COMMENT_CAT_NAMES.get(k, k), "占比": v["pct"]})
        dist_df = pd.DataFrame(dist_data)
        fig_c = px.bar(dist_df, x="类别", y="占比", color="类别",
                       color_discrete_map=COMMENT_CAT_COLORS, text_auto=".1f")
        fig_c.update_traces(texttemplate='%{text}%', textposition='outside')
        fig_c.update_layout(height=350, showlegend=False, xaxis_title="", yaxis_title="占比 (%)",
                            yaxis=dict(range=[0, max(dist_df["占比"]) * 1.3 if len(dist_df) > 0 else 30]),
                            margin=dict(t=10, b=80, l=10, r=10))
        fig_c.update_xaxes(tickangle=30)
        st.plotly_chart(fig_c, width='stretch', key="comment_bar")
    else:
        st.info("评论数据待抓取（SocialCrawl 对接中）")

# ============================================================
# Influencer Tags
# ============================================================
st.markdown("---")
st.subheader("🏷️ 达人标签")

if not df.empty:
    top_cat_name = cat_counts.index[0] if len(cat_counts) > 0 else "其他"
    if "Apple" in top_cat_name:
        topic_tag = "🍎 Apple生态型"
    elif "AI" in top_cat_name:
        topic_tag = "🤖 AI工具型"
    elif "3C" in top_cat_name:
        topic_tag = "🔌 3C种草型"
    elif "科技资讯" in top_cat_name:
        topic_tag = "📡 科技资讯型"
    elif "其他品牌" in top_cat_name:
        topic_tag = "📱 产品测评型"
    else:
        topic_tag = "🔀 混合型"

    sponsored = len(df[df["分类"].str.contains("种草|赞助|3C", na=False)])
    sp_pct = sponsored / total * 100 if total > 0 else 0
    commercial_tag = "🔴 高商业化" if sp_pct > 30 else ("🟡 中商业化" if sp_pct > 10 else "🟢 低商业化")

    score = comment_analysis.get("overall_purchase_intent_score", 0) if comment_analysis else 0
    intent_label = comment_analysis.get("overall_purchase_intent_label", "-") if comment_analysis else "-"

    tag_col1, tag_col2, tag_col3, tag_col4 = st.columns(4)
    tag_col1.markdown(f"**选题类型**: {topic_tag}")
    tag_col2.markdown(f"**商业化程度**: {commercial_tag}")
    tag_col3.markdown(f"**购买意向**: {intent_label} ({score:.2f})")
    tag_col4.markdown(f"**主要内容**: {top_cat_name[:25]}")

# ============================================================
# Video Detail List
# ============================================================
st.markdown("---")
st.subheader("📋 视频明细")

if not df.empty:
    # Filters
    f1, f2, f3 = st.columns(3)
    with f1:
        filter_cat = st.selectbox("按分类筛选", ["全部"] + sorted(df["分类"].unique().tolist()))
    with f2:
        filter_kw = st.text_input("🔎 标题关键词", placeholder="搜索...")
    with f3:
        sort_by = st.selectbox("排序", ["发布日期(新→旧)", "评论数(高→低)", "分类"])

    filtered = df.copy()
    if filter_cat != "全部":
        filtered = filtered[filtered["分类"] == filter_cat]
    if filter_kw:
        filtered = filtered[filtered["标题"].str.contains(filter_kw, case=False, na=False)]
    if sort_by == "评论数(高→低)":
        filtered = filtered.sort_values("评论数", ascending=False)
    elif sort_by == "分类":
        filtered = filtered.sort_values("分类")
    else:
        filtered = filtered.sort_values("发布日期", ascending=False)

    st.caption(f"显示 {len(filtered)} / {total} 条视频")

    comment_config = load_comment_config()

    for _, row in filtered.iterrows():
        title = row["标题"]
        title_disp = title[:150] + ("..." if len(str(title)) > 150 else "")
        cat = row["分类"]
        cat_color = CAT_COLORS.get(cat, "#F2F2F2")
        url = row.get("视频链接", "")
        comments_cnt = int(row["评论数"])

        with st.container():
            c1, c2 = st.columns([5, 1])
            with c1:
                st.markdown(
                    f'<span style="background:{cat_color};padding:3px 10px;border-radius:4px;font-size:0.85em;'
                    f'font-weight:600;margin-right:8px;">{cat}</span> '
                    f'<span style="color:#666;font-size:0.85em;">💬 {comments_cnt}</span> '
                    f'<span style="color:#999;font-size:0.8em;">{row["发布日期"]}</span>',
                    unsafe_allow_html=True
                )
                st.markdown(f"**{title_disp}**")
                if row.get("命中品牌"):
                    st.caption(f"品牌: {row['命中品牌']} | 关键词: {row.get('命中关键词', '')} | {row.get('分类依据', '')}")

            with c2:
                if url:
                    st.markdown(f'[🔗 打开视频]({url})')

                with st.expander("💬 评论明细"):
                    demo = [
                        {"text": "How much does this cost?", "likes": 89},
                        {"text": "Just ordered! Can't wait 🔥", "likes": 234},
                        {"text": "Better than the old version?", "likes": 56},
                        {"text": "Great review! Keep it up 👏", "likes": 178},
                        {"text": "link please!! 🙏", "likes": 55},
                    ]
                    for c in demo:
                        ck, _, czh, _ = classify_single_comment(c["text"], comment_config)
                        st.markdown(
                            f"👍 {c['likes']} "
                            f"<span style='background:#f0f0f0;padding:2px 6px;border-radius:3px;font-size:0.8em;'>{czh}</span> "
                            f"{c['text'][:80]}",
                            unsafe_allow_html=True
                        )

            st.markdown("---")

# ============================================================
# Comment Distribution Detail
# ============================================================
st.markdown("---")
st.subheader("💬 评论分布明细")

if comment_analysis and comment_analysis.get("aggregate_distribution"):
    cd1, cd2 = st.columns(2)
    with cd1:
        st.markdown("### 评论分类分布")
        dist_rows = []
        for k, v in sorted(comment_analysis["aggregate_distribution"].items(),
                           key=lambda x: x[1]["count"], reverse=True):
            dist_rows.append({"类别": COMMENT_CAT_NAMES.get(k, k), "占比(%)": v["pct"], "数量": v["count"]})
        st.dataframe(pd.DataFrame(dist_rows), width='stretch', hide_index=True,
                     column_config={"占比(%)": st.column_config.ProgressColumn(width="large", format="%.1f%%", min_value=0, max_value=100)})

    with cd2:
        st.markdown("### 购买意向分布")
        hi = sum(v["count"] for k, v in comment_analysis["aggregate_distribution"].items() if k in ("purchase_intent", "purchase_inquiry"))
        mi = sum(v["count"] for k, v in comment_analysis["aggregate_distribution"].items() if k in ("product_discussion", "social_engagement"))
        lo = sum(v["count"] for k, v in comment_analysis["aggregate_distribution"].items() if k in ("positive_feedback", "negative_feedback", "other"))
        tot = hi + mi + lo
        intent_df = pd.DataFrame([
            {"意向等级": "高购买意向", "占比": round(hi/tot*100, 1) if tot > 0 else 0},
            {"意向等级": "中等意向", "占比": round(mi/tot*100, 1) if tot > 0 else 0},
            {"意向等级": "低购买意向", "占比": round(lo/tot*100, 1) if tot > 0 else 0},
        ])
        fig_i = px.bar(intent_df, x="意向等级", y="占比", color="意向等级",
                       color_discrete_map={"高购买意向": "#FF6B6B", "中等意向": "#FFD93D", "低购买意向": "#6BCB77"}, text_auto=".1f")
        fig_i.update_traces(texttemplate='%{text}%', textposition='outside')
        fig_i.update_layout(height=300, showlegend=False, xaxis_title="", yaxis_title="占比 (%)",
                            yaxis=dict(range=[0, 60]), margin=dict(t=10, b=10, l=10, r=10))
        st.plotly_chart(fig_i, width='stretch', key="intent_bar")
else:
    st.info("暂无评论数据")

# ============================================================
# Download
# ============================================================
st.markdown("---")
st.subheader("📥 下载完整报告")

if not df.empty:
    wb = Workbook()

    # Sheet 1: Topics
    ws1 = wb.active
    ws1.title = "选题分类结果"
    hf = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    hfn = Font(bold=True, color="FFFFFF", size=11)
    hdrs = ["发布日期", "达人ID", "平台", "标题", "视频链接", "评论数", "分类", "命中品牌", "命中关键词", "分类依据"]
    for ci, h in enumerate(hdrs, 1):
        c = ws1.cell(row=1, column=ci, value=h)
        c.font = hfn; c.fill = hf
    for i, (_, r) in enumerate(df.iterrows(), 2):
        for j, k in enumerate(hdrs, 1):
            ws1.cell(row=i, column=j, value=r.get(k, ""))
    for col, w in [("A", 14), ("B", 18), ("C", 12), ("D", 80), ("E", 50), ("F", 10), ("G", 22), ("H", 14), ("I", 40), ("J", 35)]:
        ws1.column_dimensions[col].width = w
    ws1.freeze_panes = "A2"

    # Sheet 2: Comments
    if comment_analysis and comment_analysis.get("aggregate_distribution"):
        ws2 = wb.create_sheet("评论分析汇总")
        for ci, h in enumerate(["评论类别", "占比(%)", "数量"], 1):
            c = ws2.cell(row=1, column=ci, value=h)
            c.font = hfn; c.fill = hf
        for i, (k, v) in enumerate(sorted(comment_analysis["aggregate_distribution"].items(), key=lambda x: x[1]["count"], reverse=True), 2):
            ws2.cell(row=i, column=1, value=COMMENT_CAT_NAMES.get(k, k))
            ws2.cell(row=i, column=2, value=v["pct"])
            ws2.cell(row=i, column=3, value=v["count"])
        ws2.column_dimensions["A"].width = 20

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    st.download_button("📥 下载完整分析报告 (Excel)", data=buf.getvalue(),
                       file_name=f"达人分析_{st.session_state.username}_{datetime.now().strftime('%Y%m%d')}.xlsx",
                       mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                       use_container_width=True)

# ============================================================
# Footer
# ============================================================
st.markdown("---")
st.caption(f"v2.0 | 任意达人输入 | 缓存增量 | 中文分类 | 评论分析 | {datetime.now().strftime('%Y-%m-%d %H:%M')}")
