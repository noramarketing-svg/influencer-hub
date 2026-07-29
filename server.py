# -*- coding: utf-8 -*-
"""
达人分析系统前端 - Flask 自包含版本
参考旧版排版，支持任意达人输入
数据流: Apify(标题) + SocialCrawl(评论) → english_classifier + comment_analyzer → 前端
"""
import json, os, re, sys, time, hashlib, threading, uuid
from datetime import datetime, timedelta
from collections import Counter, defaultdict
from concurrent.futures import ThreadPoolExecutor
from flask import Flask, request, jsonify, send_from_directory
import requests
from openai import OpenAI

# 异步任务存储（内存字典）
_async_tasks = {}  # {task_id: {status, result, error, ...}}

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, os.path.join(BASE_DIR, "configs"))
import api_keys as _api_keys

def reload_api_keys():
    """重新加载 configs/api_keys.py，支持多 Key（逗号分隔）"""
    import importlib
    importlib.reload(_api_keys)
    global SC_API_KEYS, SC_API_KEY, SC_HEADERS, DEEPSEEK_API_KEY, DEEPSEEK_CLIENT, APIFY_API_KEY
    # ScrapeCreators 支持多 Key：取第一个作为当前 active
    raw_sc = getattr(_api_keys, "SCRAPECREATORS_API_KEY", "")
    SC_API_KEYS = [k.strip() for k in raw_sc.split(",") if k.strip()]
    SC_API_KEY = SC_API_KEYS[0] if SC_API_KEYS else ""
    SC_HEADERS = {"x-api-key": SC_API_KEY}
    DEEPSEEK_API_KEY = getattr(_api_keys, "DEEPSEEK_API_KEY", "")
    DEEPSEEK_CLIENT = OpenAI(api_key=DEEPSEEK_API_KEY, base_url="https://api.deepseek.com") if DEEPSEEK_API_KEY else None
    APIFY_API_KEY = getattr(_api_keys, "APIFY_API_KEY", "")
    try:
        import socialcrawl_fetcher
        socialcrawl_fetcher.SCRAPECREATORS_API_KEY = SC_API_KEY
    except Exception:
        pass

reload_api_keys()

app = Flask(__name__)

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, os.path.join(BASE_DIR, "engines"))

from english_classifier import classify_title_en, load_config as load_en_config
from spanish_classifier import classify_title_es, load_config as load_es_config
from comment_analyzer import classify_single_comment, analyze_video_comments, analyze_account_comments, load_config as load_comment_config
from socialcrawl_fetcher import fetch_ig_comments_sc, fetch_tiktok_comments_sc, get_top_valid_comments, fetch_comments_for_video
from apify_fetcher import fetch_instagram_videos, fetch_tiktok_videos

CAT_COLORS = {
    "3C配件品牌赞助/种草": "#F8CBAD",
    "Apple/iOS生态": "#DDEBF7",
    "其他品牌产品种草": "#FCE4D6",
    "科技资讯/教程技巧": "#D9E1F2",
    "AI工具/生活观点": "#E2EFDA",
    "赞助广告内容": "#F4CCCC",
    "其他": "#F2F2F2",
}

COMMENT_CAT_NAMES = {
    "content_engagement": "内容互动",
    "purchase_intent": "购买意向",
    "product_interaction": "产品互动",
    "other": "其他",
}

COMMENT_CAT_CSS = {
    "content_engagement": "c-engagement",
    "purchase_intent": "c-intent",
    "product_interaction": "c-discuss",
    "other": "",
}

SPONSORED_CATEGORIES = ["3C配件品牌赞助/种草", "其他品牌产品种草", "赞助广告内容"]

# ============================================================
# 达人资源总库 — 索引 + 增量更新
# ============================================================
CACHE_DIR = os.path.join(BASE_DIR, "cache")
INDEX_PATH = os.path.join(CACHE_DIR, "influencer_index.json")

def load_influencer_index():
    """加载达人资源总库索引"""
    if os.path.exists(INDEX_PATH):
        with open(INDEX_PATH, "r", encoding="utf-8") as f:
            return json.load(f)
    return {"instagram": {}, "tiktok": {}}

def save_influencer_index(index):
    """保存达人资源总库索引"""
    os.makedirs(CACHE_DIR, exist_ok=True)
    with open(INDEX_PATH, "w", encoding="utf-8") as f:
        json.dump(index, f, ensure_ascii=False, indent=2)

def update_index(username, platform, videos):
    """更新达人索引"""
    index = load_influencer_index()
    pf = platform.lower()
    if pf not in index:
        index[pf] = {}
    
    dates = [v.get("发布日期", "") for v in videos if v.get("发布日期")]
    dates.sort()
    index[pf][username] = {
        "last_fetch": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "video_count": len(videos),
        "newest_video_date": dates[-1] if dates else "",
        "oldest_video_date": dates[0] if dates else "",
    }
    save_influencer_index(index)

def merge_videos(cached, new):
    """按视频链接去重合并，新数据覆盖旧数据"""
    merged = {v.get("视频链接", v.get("url", "")): v for v in cached}
    for v in new:
        key = v.get("视频链接", v.get("url", ""))
        if key:
            merged[key] = v  # 新数据覆盖
    result = list(merged.values())
    result.sort(key=lambda v: v.get("发布日期", ""), reverse=True)
    return result

def needs_apify_refresh(username, platform, days):
    """
    判断是否需要 Apify 补抓
    返回: (need_refresh, cached_videos, message)
    """
    cached = load_cached_videos(username, platform)
    if not cached:
        return (True, [], "无历史缓存，需全量抓取")
    
    dates = [v.get("发布日期", "") for v in cached if v.get("发布日期")]
    if not dates:
        return (True, cached, "缓存数据无日期，需重新抓取")
    
    newest = max(dates)
    today = datetime.now().date()
    cutoff = (today - timedelta(days=max(days - 1, 0))).strftime("%Y-%m-%d")
    fresh_through = today.strftime("%Y-%m-%d")
    filtered = [v for v in cached if v.get("发布日期", "") >= cutoff]

    # 是否抓取过与“最新视频发布日期”是两件事：达人今天可能没有发视频。
    index = load_influencer_index()
    account_meta = index.get(platform.lower(), {}).get(username, {})
    last_fetch = str(account_meta.get("last_fetch", ""))
    fetched_today = last_fetch[:10] == fresh_through

    if fetched_today:
        return (
            False,
            filtered,
            f"✅ 今日已更新（抓取时间 {last_fetch}，最新视频 {newest}），共 {len(filtered)} 条",
        )
    else:
        return (
            True,
            filtered,
            f"⚠️ 缓存最新视频 {newest}，上次抓取 {last_fetch or '未知'}，需补抓至 {fresh_through}",
        )


def filter_videos_by_days(videos, days):
    """保留包含今天在内的近 N 个自然日数据。"""
    cutoff = (datetime.now().date() - timedelta(days=max(int(days) - 1, 0))).strftime("%Y-%m-%d")
    return [v for v in videos if not v.get("发布日期") or v.get("发布日期", "") >= cutoff]

def load_cached_videos(username, platform):
    cache_dir = os.path.join(CACHE_DIR, platform.lower())
    os.makedirs(cache_dir, exist_ok=True)
    cache_file = os.path.join(cache_dir, f"{username}_videos.json")
    if os.path.exists(cache_file):
        with open(cache_file, "r", encoding="utf-8") as f:
            return json.load(f)
    return []

def save_cached_videos(username, platform, videos):
    cache_dir = os.path.join(CACHE_DIR, platform.lower())
    os.makedirs(cache_dir, exist_ok=True)
    cache_file = os.path.join(cache_dir, f"{username}_videos.json")
    with open(cache_file, "w", encoding="utf-8") as f:
        json.dump(videos, f, ensure_ascii=False, indent=2)
    update_index(username, platform, videos)


def extract_username(text, platform):
    text = text.strip()
    if "tiktok.com" in text:
        m = re.search(r'@([a-zA-Z0-9_.]+)', text)
        return m.group(1) if m else text.split("@")[-1].split("/")[0].split("?")[0]
    if "instagram.com" in text:
        m = re.search(r'instagram\.com/([a-zA-Z0-9_.]+)', text)
        return m.group(1) if m else text.split("/")[-1].split("?")[0]
    return text.replace("@", "").split("/")[0].split("?")[0].strip()


@app.route("/health")
def health():
    """Health check + version info"""
    return jsonify({
        "status": "ok",
        "version": "4.0",
        "features": ["batch_analyze", "batch_fetch", "comment_stop", "batch_stop", "llm_fallback", "panel4_batch_comments"]
    })


@app.route("/downloads/<path:filename>")
def serve_download(filename):
    """提供回填 Excel 下载"""
    download_dir = os.path.join(BASE_DIR, "downloads")
    return send_from_directory(download_dir, filename, as_attachment=True)


@app.route("/")
def index():
    """Serve the main page"""
    html_path = os.path.join(BASE_DIR, "templates", "index.html")
    if os.path.exists(html_path):
        with open(html_path, "r", encoding="utf-8") as f:
            return f.read()
    return "<h1>Template not found</h1>"


@app.route("/report")
def api_report():
    """Serve the API evaluation report"""
    html_path = os.path.join(BASE_DIR, "templates", "api-report.html")
    if os.path.exists(html_path):
        with open(html_path, "r", encoding="utf-8") as f:
            return f.read()
    return "<h1>Report not found</h1>"


# ============================================================
# API: Analyze influencer — 标题抓取 + 分类
# ============================================================
@app.route("/api/analyze", methods=["POST"])
def api_analyze():
    """Analyze an influencer - 增量更新 + 分类"""
    try:
        data = request.get_json()
        input_text = data.get("input", "")
        platform = data.get("platform", "TikTok")
        days = int(data.get("days", 30))
        language = data.get("language", "en")

        username = extract_username(input_text, platform)
        if not username:
            return jsonify({"error": "无法解析达人ID"}), 400

        en_config = load_en_config() if language != "es" else None
        es_config = load_es_config() if language == "es" else None
        classify_fn = lambda t: _wrap_classify(t, language, en_config, es_config)

        # 增量更新判断
        need_refresh, cached, status_msg = needs_apify_refresh(username, platform, days)

        if not need_refresh:
            # 缓存最新，直接用
            results = []
            unclassified_titles = []
            unclassified_indices = []
            for i, v in enumerate(cached):
                title = v.get("标题", v.get("title", ""))
                brand, keywords, category, basis = classify_fn(title)
                if category == "其他" and basis == "No rule matched":
                    unclassified_titles.append(title)
                    unclassified_indices.append(i)
                results.append({
                    "发布日期": v.get("发布日期", v.get("date", "")),
                    "达人ID": username,
                    "平台": platform,
                    "标题": title,
                    "视频链接": v.get("视频链接", v.get("url", "")),
                    "评论数": int(v.get("评论数", v.get("comments_count", 0)) or 0),
                    "分类": category,
                    "命中品牌": brand or "",
                    "命中关键词": keywords or "",
                    "分类依据": basis,
                })
            # LLM 兜底
            if language != "es" and unclassified_titles:
                en_config = load_en_config()
                llm_results = _llm_classify_titles(unclassified_titles, en_config)
                for local_idx, cat in llm_results.items():
                    if local_idx < len(unclassified_indices):
                        result_idx = unclassified_indices[local_idx]
                        results[result_idx]["分类"] = cat
                        results[result_idx]["分类依据"] = "LLM fallback"
            return jsonify({
                "username": username, "platform": platform, "language": language,
                "videos": results, "source": "cache", "status": status_msg
            })

        # 需要补抓 — 返回状态提示，前端触发 Apify
        if cached:
            return jsonify({
                "username": username, "platform": platform, "language": language,
                "videos": [],
                "source": "stale",
                "need_apify": True,
                "cached_count": len(cached),
                "message": status_msg + "，请点击下方按钮抓取最新数据。"
            })
        else:
            return jsonify({
                "username": username, "platform": platform, "language": language,
                "videos": [],
                "source": "empty",
                "need_apify": True,
                "cached_count": 0,
                "message": "无缓存数据，请点击下方按钮抓取。"
            })
    except Exception as e:
        import traceback
        return jsonify({"error": f"分析失败: {str(e)}", "trace": traceback.format_exc()}), 500


# ============================================================
# API: Single video comments (备用)
# ============================================================
@app.route("/api/comments", methods=["POST"])
def api_comments():
    """Fetch real comments from SocialCrawl for a single video"""
    data = request.get_json()
    video_url = data.get("url", "")
    platform = data.get("platform", "")

    try:
        if "instagram" in platform.lower():
            comments = fetch_ig_comments_sc(video_url)
        elif "tiktok" in platform.lower():
            comments = fetch_tiktok_comments_sc(video_url)
        else:
            return jsonify({"error": "Unknown platform"}), 400

        valid = get_top_valid_comments(comments, 30)
        comment_config = load_comment_config()

        results = []
        for c in valid:
            cat_key, name_en, name_zh, signals = classify_single_comment(c["text"], comment_config)
            results.append({
                "text": c["text"][:200],
                "likes": c["likes"],
                "username": c.get("username", ""),
                "category": name_zh,
                "category_key": cat_key,
                "matched_signals": signals,
            })

        return jsonify({"comments": results, "total": len(comments)})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ============================================================
# API: Batch comment analysis — 两个 Part
# ============================================================
@app.route("/api/comments/analysis", methods=["POST"])
def api_comments_analysis():
    """
    一次性获取两个 Part 的评论分析：
    - Part 1: TOP-3 评论数最高的视频
    - Part 2: TOP-3 赞助/种草分类中评论数最高的视频

    Input: {videos: [{...}], platform: "Instagram|TikTok"}
    Output: {part1_top3_hot: {...}, part2_top3_sponsored: {...}}
    """
    data = request.get_json()
    videos = data.get("videos", [])
    platform = data.get("platform", "Instagram")

    if not videos:
        print(f"[comments/analysis] ERROR: videos is empty. data keys: {list(data.keys()) if data else 'None'}, videos type: {type(data.get('videos'))}")
        return jsonify({"error": "No videos provided"}), 400

    comment_config = load_comment_config()

    # --- Part 1: TOP-3 by comment count ---
    sorted_by_comments = sorted(videos, key=lambda v: int(v.get("评论数", 0) or 0), reverse=True)
    top3_hot = sorted_by_comments[:3]

    # --- Part 2: TOP-3 sponsored/review by comment count ---
    sponsored_videos = [v for v in videos if v.get("分类", "") in SPONSORED_CATEGORIES]
    # If < 3 sponsored videos, fill with next-highest comment count non-sponsored
    if len(sponsored_videos) < 3:
        non_sponsored = [v for v in sorted_by_comments if v not in sponsored_videos]
        needed = 3 - len(sponsored_videos)
        sponsored_videos.extend(non_sponsored[:needed])
    top3_sponsored = sorted(sponsored_videos, key=lambda v: int(v.get("评论数", 0) or 0), reverse=True)[:3]

    # --- Fetch & classify comments for each video (parallel for speed) ---
    def analyze_single_video(v):
        """Fetch + classify comments for ONE video. Runs inside a worker thread."""
        url = v.get("视频链接", "")
        title = v.get("标题", "")
        native_comment_count = int(v.get("评论数", 0) or 0)

        # Fetch comments from ScrapeCreators
        comments_raw = []
        fetch_error = None
        if url:
            try:
                comments_raw = fetch_comments_for_video(url, platform)
            except Exception as e:
                fetch_error = str(e)

        valid_comments = get_top_valid_comments(comments_raw, 30)

        # Classify each comment
        classified = []
        for c in valid_comments:
            cat_key, name_en, name_zh, signals = classify_single_comment(c["text"], comment_config)
            classified.append({
                "text": c["text"][:200],
                "likes": c["likes"],
                "username": c.get("username", "N/A"),
                "category": name_zh,
                "category_key": cat_key,
                "matched_signals": signals,
            })

        # Distribution for this video
        dist = defaultdict(lambda: {"count": 0, "pct": 0})
        for c in classified:
            dist[c["category_key"]]["count"] += 1
        total_c = len(classified)
        for k in dist:
            dist[k]["pct"] = round(dist[k]["count"] / total_c * 100, 1) if total_c > 0 else 0

        return {
            "title": title[:120],
            "url": url,
            "native_comment_count": native_comment_count,
            "fetched_comment_count": len(classified),
            "total_raw_comments": len(comments_raw),
            "distribution": dict(dist),
            "classified_comments": classified,
            "fetch_error": fetch_error,
            "_classified": classified,  # internal: for group aggregation
        }

    def _empty_group():
        return {
            "videos": [],
            "aggregate_distribution": {},
            "total_classified_comments": 0,
            "purchase_intent_score": 0,
            "purchase_intent_label": "低购买意向",
        }

    def analyze_video_group(video_group, analyzed_map):
        """Assemble a group's result from the pre-fetched analyzed map."""
        if not video_group:
            return _empty_group()

        video_results = []
        all_classified = []
        # Track unclassified for LLM fallback
        unclassified_texts = []
        unclassified_global_indices = []
        
        for v in video_group:
            res = analyzed_map.get(v.get("视频链接", ""))
            if res is None:
                continue
            for c in res.get("_classified", []):
                all_classified.append(c)
                if c["category_key"] == "other":
                    unclassified_texts.append(c.get("text", "")[:200])
                    unclassified_global_indices.append(len(all_classified) - 1)
            video_results.append({k: val for k, val in res.items() if k != "_classified"})

        # LLM 兜底 — 对 "other" 评论做语义重分类
        if comment_config.get("llm_fallback", {}).get("enabled") and unclassified_texts:
            llm_results = _llm_classify_comments(unclassified_texts, comment_config)
            for local_idx, cat_key in llm_results.items():
                if local_idx < len(unclassified_global_indices):
                    g_idx = unclassified_global_indices[local_idx]
                    all_classified[g_idx]["category_key"] = cat_key
                    all_classified[g_idx]["category"] = COMMENT_CAT_NAMES.get(cat_key, "其他")

        # Aggregate distribution across all videos in group
        agg_dist = defaultdict(lambda: {"count": 0, "name_zh": "", "pct": 0})
        for c in all_classified:
            cat_key = c["category_key"]
            agg_dist[cat_key]["count"] += 1
            agg_dist[cat_key]["name_zh"] = c["category"]

        total_all = len(all_classified)
        for k in agg_dist:
            agg_dist[k]["pct"] = round(agg_dist[k]["count"] / total_all * 100, 1) if total_all > 0 else 0

        # Purchase intent scoring
        intent_score, intent_label = _calc_purchase_intent(all_classified, comment_config)

        return {
            "videos": video_results,
            "aggregate_distribution": dict(agg_dist),
            "total_classified_comments": total_all,
            "purchase_intent_score": intent_score,
            "purchase_intent_label": intent_label,
        }

    # --- Fetch ALL unique videos ONCE in parallel (caps total time at ~1 video) ---
    unique_videos = []
    _seen = set()
    for v in (top3_hot + top3_sponsored):
        u = v.get("视频链接", "")
        if u and u not in _seen:
            _seen.add(u)
            unique_videos.append(v)

    analyzed_map = {}
    if unique_videos:
        workers = min(6, len(unique_videos))
        with ThreadPoolExecutor(max_workers=workers) as ex:
            futures = {ex.submit(analyze_single_video, v): v.get("视频链接", "") for v in unique_videos}
            for f in futures:
                analyzed_map[futures[f]] = f.result()

    part1 = analyze_video_group(top3_hot, analyzed_map)
    part2 = analyze_video_group(top3_sponsored, analyzed_map)

    return jsonify({
        "part1_top3_hot": part1,
        "part2_top3_sponsored": part2,
        "platform": platform,
    })


def _calc_purchase_intent(classified_comments, config):
    """Calculate purchase intent score from classified comments"""
    scoring = config.get("purchase_intent_scoring", {})
    weights = scoring.get("weights", {})
    interpretation = scoring.get("interpretation", {})

    if not classified_comments:
        return 0, "低购买意向"

    total_weighted = sum(weights.get(c["category_key"], 0) for c in classified_comments)
    avg_score = total_weighted / len(classified_comments)

    if avg_score >= interpretation.get("high", {}).get("min_score", 2.0):
        label = interpretation["high"].get("label_zh", "高购买意向")
    elif avg_score >= interpretation.get("medium", {}).get("min_score", 1.0):
        label = interpretation["medium"].get("label_zh", "中购买意向")
    else:
        label = interpretation["low"].get("label_zh", "低购买意向")

    return round(avg_score, 2), label


# ============================================================
# Helpers: 单条视频分析（板块 2）+ 批量 View 更新（板块 3）
# ============================================================
def _normalize_url(url):
    """标准化视频链接：处理 Instagram /reels/ → /reel/，去除 query 参数"""
    if not url:
        return url
    url = url.strip()
    # Instagram /reels/ 统一为 /reel/
    url = re.sub(r'instagram\.com/reels/', 'instagram.com/reel/', url, flags=re.IGNORECASE)
    # 去除 query 参数
    if '?' in url:
        url = url.split('?')[0]
    return url


def _raise_sc_error(response, platform_label, url):
    """根据 HTTP 状态码抛出区分度更好的中文错误"""
    status = response.status_code
    try:
        body = response.json()
        msg = body.get("message", response.text[:200])
    except Exception:
        msg = response.text[:200]

    if status == 402:
        raise Exception(f"SCRAPECREATORS_CREDIT_EXHAUSTED: {platform_label} API 额度已用完，请在 API 设置中更换 ScrapeCreators Key。原始信息：{msg}")
    elif status == 404:
        raise Exception(f"SCRAPECREATORS_404: 该 {platform_label} 视频无法访问（可能为私密账号、已删除或地区限制）。请确认链接可正常打开。链接：{url}")
    elif status == 429:
        raise Exception(f"SCRAPECREATORS_RATE_LIMITED: {platform_label} API 请求过于频繁，请稍后重试。原始信息：{msg}")
    elif status == 401:
        raise Exception(f"SCRAPECREATORS_AUTH_FAILED: {platform_label} API Key 无效，请在 API 设置中检查。原始信息：{msg}")
    else:
        raise Exception(f"{platform_label} API 请求失败 (HTTP {status}): {msg}")


def _detect_platform(url):
    """从 URL 识别平台"""
    url = (url or "").lower()
    if "tiktok.com" in url:
        return "TikTok"
    if "instagram.com" in url:
        return "Instagram"
    return ""


# ============ 多 Key 轮换 + 额度查询 ============

# 记录当前 active 的 ScrapeCreators key 索引
_sc_key_index = 0
_sc_key_lock = threading.Lock()

def _get_active_sc_key():
    global _sc_key_index
    with _sc_key_lock:
        if not SC_API_KEYS:
            return ""
        return SC_API_KEYS[_sc_key_index % len(SC_API_KEYS)]

def _rotate_sc_key():
    global _sc_key_index
    with _sc_key_lock:
        if len(SC_API_KEYS) <= 1:
            return False
        _sc_key_index = (_sc_key_index + 1) % len(SC_API_KEYS)
    return True


def _mask_key(key):
    """遮罩 API Key，供状态接口安全展示。"""
    if not key:
        return "未设置"
    if len(key) <= 8:
        return key[:2] + "***"
    return key[:4] + "***" + key[-4:]

def _sc_request(method, url, **kwargs):
    """带自动轮换的 ScrapeCreators 请求：402 时自动切下一个 key"""
    timeout = kwargs.pop("timeout", 30)
    for attempt in range(len(SC_API_KEYS)):
        key = _get_active_sc_key()
        headers = {"x-api-key": key}
        try:
            r = requests.request(method, url, headers=headers, timeout=timeout, **kwargs)
            if r.status_code == 402:
                if not _rotate_sc_key():
                    break
                continue
            return r
        except requests.exceptions.RequestException:
            if attempt + 1 >= len(SC_API_KEYS):
                raise
            _rotate_sc_key()
    return requests.request(method, url, headers={"x-api-key": SC_API_KEYS[0]}, timeout=timeout, **kwargs)

def _check_sc_credits():
    """查询 ScrapeCreators 各 Key 的状态（通过轻量测试请求判断）"""
    if not SC_API_KEYS:
        return {"keys": [], "active": 0}
    keys_status = []
    active_idx = 0
    for i, key in enumerate(SC_API_KEYS):
        try:
            r = requests.get("https://api.scrapecreators.com/v2/tiktok/video",
                             params={"url": "https://www.tiktok.com/@test/video/1"},
                             headers={"x-api-key": key}, timeout=10)
            if r.status_code == 402:
                keys_status.append({"index": i, "masked": _mask_key(key), "status": "exhausted"})
            elif r.status_code in (200, 400, 404, 422):
                keys_status.append({"index": i, "masked": _mask_key(key), "status": "active"})
                if i == _sc_key_index % len(SC_API_KEYS):
                    active_idx = i
            elif r.status_code in (401, 403):
                keys_status.append({"index": i, "masked": _mask_key(key), "status": "invalid"})
            elif r.status_code == 429:
                keys_status.append({"index": i, "masked": _mask_key(key), "status": "limited"})
            else:
                keys_status.append({"index": i, "masked": _mask_key(key), "status": f"error_{r.status_code}"})
        except Exception as e:
            keys_status.append({"index": i, "masked": _mask_key(key), "status": f"error: {str(e)[:30]}"})
    return {"keys": keys_status, "active": active_idx, "total": len(SC_API_KEYS)}


def _check_deepseek_status():
    """DeepSeek 不提供通用余额查询；用模型列表接口判断 Key 是否可用。"""
    if not DEEPSEEK_API_KEY:
        return {"status": "missing"}
    try:
        r = requests.get(
            "https://api.deepseek.com/models",
            headers={"Authorization": f"Bearer {DEEPSEEK_API_KEY}"},
            timeout=10,
        )
        if r.status_code == 200:
            return {"status": "active"}
        if r.status_code == 402:
            return {"status": "exhausted"}
        if r.status_code in (401, 403):
            return {"status": "invalid"}
        if r.status_code == 429:
            return {"status": "limited"}
        return {"status": "unavailable", "http_status": r.status_code}
    except Exception as e:
        return {"status": "unavailable", "error": str(e)[:100]}

def _check_apify_credits():
    """查询 Apify 剩余额度"""
    if not APIFY_API_KEY:
        return None
    try:
        r = requests.get("https://api.apify.com/v2/users/me",
                         headers={"Authorization": f"Bearer {APIFY_API_KEY}"}, timeout=10)
        if r.status_code == 200:
            try:
                data = r.json()
                uid = data.get("data", {}).get("id")
            except Exception:
                return {"remaining": "?", "error": "Apify response not JSON", "raw_status": r.status_code}
            if uid:
                r2 = requests.get(f"https://api.apify.com/v2/users/{uid}/usage/monthly",
                                  headers={"Authorization": f"Bearer {APIFY_API_KEY}"}, timeout=10)
                if r2.status_code == 200:
                    try:
                        u = r2.json().get("data", {})
                    except Exception:
                        return {"remaining": "?", "error": "Apify usage response not JSON"}
                    used = u.get("totalUsageCreditsUsdAfterVolumeDiscount", 0)
                    free_tier = 5.0
                    remaining = max(0, free_tier - used)
                    return {"remaining": round(remaining, 2), "used": round(used, 2), "total": free_tier}
                else:
                    return {"remaining": "?", "error": f"Apify usage HTTP {r2.status_code}"}
            else:
                return {"remaining": "?", "error": "No uid in Apify response"}
        else:
            return {"remaining": "?", "error": f"Apify HTTP {r.status_code}"}
    except Exception as e:
        return {"remaining": "?", "error": str(e)[:80]}


def _fetch_video_metadata_sc(url, platform):
    """通过 ScrapeCreators 获取单视频基础数据（自动轮换 Key）"""
    platform_lower = (platform or "").lower()
    if "tiktok" in platform_lower:
        r = _sc_request("GET", "https://api.scrapecreators.com/v2/tiktok/video", params={"url": url}, timeout=30)
        if r.status_code != 200:
            _raise_sc_error(r, "TikTok", url)
        data = r.json()
        ad = data.get("aweme_detail", {})
        stats = ad.get("statistics", {})
        author = ad.get("author", {})
        return {
            "platform": "TikTok",
            "video_id": stats.get("aweme_id", ""),
            "url": url,
            "author": author.get("unique_id", ""),
            "title": ad.get("desc", ""),
            "thumbnail": ad.get("cover", {}).get("url_list", [""])[0] if ad.get("cover") else "",
            "views": stats.get("play_count", 0),
            "likes": stats.get("digg_count", 0),
            "comments": stats.get("comment_count", 0),
            "shares": stats.get("share_count", 0),
            "collects": stats.get("collect_count", 0),
            "published_at": ad.get("create_time", ""),
            "duration": ad.get("duration", 0),
        }
    elif "instagram" in platform_lower:
        r = _sc_request("GET", "https://api.scrapecreators.com/v1/instagram/post", params={"url": url}, timeout=30)
        if r.status_code != 200:
            _raise_sc_error(r, "Instagram", url)
        data = r.json()
        media = data.get("data", {}).get("xdt_shortcode_media", {})
        owner = media.get("owner", {}) or {}

        # Caption: try direct caption first, fallback to edge_media_to_caption
        caption = ""
        if media.get("caption"):
            caption = media["caption"].get("text", "") or ""
        if not caption:
            edge_caption = media.get("edge_media_to_caption", {})
            edges = edge_caption.get("edges", [])
            if edges:
                caption = edges[0].get("node", {}).get("text", "") or ""

        # Views: video_play_count is more accurate for Reels
        views = media.get("video_play_count", 0) or media.get("video_view_count", 0) or 0

        # Likes: like_count is often N/A, use edge_media_preview_like
        likes = media.get("like_count", 0)
        if not likes:
            likes = media.get("edge_media_preview_like", {}).get("count", 0) or 0

        # Comments: comment_count is often N/A, use edge counts
        comments = media.get("comment_count", 0)
        if not comments:
            comments = media.get("edge_media_to_parent_comment", {}).get("count", 0) or 0
            if not comments:
                comments = media.get("edge_media_preview_comment", {}).get("count", 0) or 0

        return {
            "platform": "Instagram",
            "video_id": media.get("shortcode", ""),
            "url": url,
            "author": owner.get("username", ""),
            "title": caption,
            "thumbnail": media.get("thumbnail_src", ""),
            "views": views,
            "likes": likes,
            "comments": comments,
            "shares": None,
            "collects": None,
            "published_at": media.get("taken_at_timestamp", ""),
            "duration": media.get("video_duration", 0),
        }
    else:
        raise Exception(f"Unsupported platform: {platform}")


def _fetch_transcript_sc(url, platform):
    """通过 ScrapeCreators 获取视频口播原文（自动轮换 Key）"""
    platform_lower = (platform or "").lower()
    if "tiktok" in platform_lower:
        r = _sc_request("GET", "https://api.scrapecreators.com/v1/tiktok/video/transcript", params={"url": url, "language": "en"}, timeout=45)
        if r.status_code != 200:
            return None
        data = r.json()
        transcript = data.get("transcript")
        if not transcript:
            return None
        # WEBVTT 格式，简单清理
        lines = []
        for line in transcript.splitlines():
            line = line.strip()
            if not line or line.startswith("WEBVTT") or "-->" in line or line.isdigit():
                continue
            lines.append(line)
        return " ".join(lines)
    elif "instagram" in platform_lower:
        r = _sc_request("GET", "https://api.scrapecreators.com/v2/instagram/media/transcript", params={"url": url}, timeout=60)
        if r.status_code != 200:
            return None
        data = r.json()
        transcripts = data.get("transcripts", [])
        if transcripts and transcripts[0].get("text"):
            return transcripts[0].get("text")
        return None
    return None


def _is_single_video_url(url, platform):
    """校验是否为单条视频，而非达人主页或 Reels 列表页。"""
    clean = (url or "").lower().split("?")[0].rstrip("/") + "/"
    platform_lower = (platform or "").lower()
    if "instagram" in platform_lower:
        return bool(re.search(r"instagram\.com/(reel|p|tv)/[^/]+/", clean))
    if "tiktok" in platform_lower:
        return bool(re.search(r"tiktok\.com/@[^/]+/video/\d+/", clean))
    return False


def _fetch_comments_single_video_sc(url, platform, target_valid=50):
    """抓取单条视频的评论，最多 target_valid 条有效评论"""
    platform_lower = (platform or "").lower()
    fetch_limit = 10000 if target_valid is None else max(int(target_valid), 0)
    if fetch_limit == 0:
        return []
    max_pages = 100 if target_valid is None else 15
    if "instagram" in platform_lower:
        comments = fetch_ig_comments_sc(url, target_valid=fetch_limit, max_pages=max_pages)
    elif "tiktok" in platform_lower:
        comments = fetch_tiktok_comments_sc(url, target_valid=fetch_limit, max_pages=max_pages)
    else:
        comments = []
    # 按点赞排序，保留有效评论
    valid = [c for c in comments if c.get("valid", True)]
    return valid if target_valid is None else valid[:fetch_limit]


def _translate_full_text_to_chinese(text):
    """将口播全文翻译为中文，保持段落和信息完整。"""
    if not text:
        return ""
    if not DEEPSEEK_CLIENT:
        return "（未配置 DeepSeek，无法生成中文翻译）"
    try:
        response = DEEPSEEK_CLIENT.chat.completions.create(
            model="deepseek-chat",
            messages=[
                {"role": "system", "content": "你是专业翻译。把用户提供的口播逐段完整翻译成简体中文，不总结、不删减，只输出译文。"},
                {"role": "user", "content": text},
            ],
            max_tokens=4000,
            temperature=0.2,
        )
        return response.choices[0].message.content.strip()
    except Exception as e:
        return f"（中文翻译失败：{str(e)[:120]}）"


def _analyze_video_with_deepseek(transcript, comments_sample, language="en"):
    """调用 DeepSeek 分析视频口播：HOOK、中段、CTA"""
    if not DEEPSEEK_CLIENT:
        return {"error": "DeepSeek 未配置"}

    comments_text = "\n".join([f"- {c.get('text','')}" for c in comments_sample[:15]])
    prompt = f"""你是一位资深短视频内容分析专家。请基于以下视频口播原文和代表性评论，对这条视频进行结构化拆解。

【口播原文】
{transcript or '（未获取到口播原文）'}

【代表性评论】
{comments_text or '（无评论）'}

请按以下 JSON 格式输出（只输出 JSON，不要多余说明）：
{{
  "hook": "前3秒如何吸引用户注意，具体话术/画面逻辑",
  "middle": "中段核心内容/卖点/节奏分析",
  "cta": "结尾引导行为/转化话术分析",
  "content_summary": "整条视频内容一句话总结",
  "audience_insight": "从评论中看出观众最关注什么",
  "improvement_suggestions": "可优化建议（3-5条）"
}}
"""
    try:
        r = DEEPSEEK_CLIENT.chat.completions.create(
            model="deepseek-chat",
            messages=[
                {"role": "system", "content": "你是短视频内容分析专家，只输出结构化的 JSON。"},
                {"role": "user", "content": prompt}
            ],
            max_tokens=1500,
            temperature=0.7,
        )
        content = r.choices[0].message.content
        # 尝试解析 JSON
        try:
            # 清理可能的 markdown 代码块
            content = content.strip()
            if content.startswith("```"):
                content = re.sub(r'^```(?:json)?\s*', '', content)
                content = re.sub(r'\s*```$', '', content)
            return json.loads(content)
        except json.JSONDecodeError:
            return {"raw_response": content, "hook": content[:200]}
    except Exception as e:
        return {"error": f"DeepSeek API 调用失败: {str(e)}"}


# ============================================================
# LLM Fallback: 分类兜底（选题 + 评论）
# ============================================================

def _llm_classify_titles(titles, config):
    """LLM 兜底：对未命中的视频标题做语义分类（批量，最多 20 条）"""
    if not DEEPSEEK_CLIENT:
        return {}
    
    llm_cfg = config.get("llm_config", {})
    if not llm_cfg.get("enabled", False):
        return {}
    
    prompt_template = llm_cfg.get("prompt_template", "")
    if not prompt_template:
        prompt_template = """你是一个短视频标题分类器。请将以下每个标题分入最合适的类别。
类别选项:
1. 3C配件品牌赞助/种草 (3C Accessories Brand Sponsor/Review) - 对3C配件(充电器/耳机/手机壳等)的深度评测、对比或推荐
2. Apple/iOS生态 (Apple/iOS Ecosystem) - 苹果产品、iOS、Mac生态相关内容
3. 其他品牌产品种草 (Other Brand Product Review) - 对其他品牌产品的评测或推荐
4. 科技资讯/教程技巧 (Tech News & Tutorials) - 科技新闻、教程、产品发布、技术科普
5. AI工具/生活观点 (AI Tools & Tech Lifestyle) - AI工具使用、科技生活方式
6. 赞助广告内容 (Sponsored Content)
7. 其他 (Other) - 无法归类的非技术内容

对每条标题，输出JSON格式：{"index": 序号, "category": "类别名", "reason": "简短理由"}

标题列表：
"""
    titles_text = "\n".join([f"{i}. {t}" for i, t in enumerate(titles)])
    full_prompt = prompt_template + titles_text + "\n\n只输出JSON数组，不要其他内容。"
    
    try:
        r = DEEPSEEK_CLIENT.chat.completions.create(
            model="deepseek-chat",
            messages=[
                {"role": "system", "content": "你是短视频标题分类专家，只输出结构化的JSON数组。"},
                {"role": "user", "content": full_prompt}
            ],
            max_tokens=2000,
            temperature=0.3,
        )
        content = r.choices[0].message.content.strip()
        if content.startswith("```"):
            content = re.sub(r'^```(?:json)?\s*', '', content)
            content = re.sub(r'\s*```$', '', content)
        results = json.loads(content)
        if isinstance(results, list):
            return {item.get("index", 0): item.get("category", "其他") for item in results}
        return {}
    except Exception as e:
        print(f"[LLM fallback] Title classification failed: {e}")
        return {}


def _llm_classify_comments(comments, config):
    """LLM 兜底：对未命中的评论做语义分类（批量，最多 30 条）"""
    if not DEEPSEEK_CLIENT:
        return {}
    
    llm_cfg = config.get("llm_fallback", {})
    if not llm_cfg.get("enabled", False):
        return {}
    
    prompt_template = llm_cfg.get("prompt_template", "")
    if not prompt_template:
        prompt_template = """你是一个TikTok/Instagram评论分类器。请将以下每条评论分入最合适的类别。
类别选项:
1. content_engagement (内容互动) - 对视频内容的反应、表情、笑话、对创作者的赞美
2. purchase_intent (购买意向) - 对产品的购买意愿：询价、问购买渠道、要折扣码、表达想要的意愿
3. product_interaction (产品互动) - 讨论产品本身：问功能、对比品牌、分享使用体验、给建议
4. other (其他) - 无法归类的评论：@好友、纯表情、与产品/内容无关的闲聊

对每条评论，输出JSON格式：{"index": 序号, "category": "类别key", "reason": "简短理由"}

评论列表：
"""
    comments_text = "\n".join([f"{i}. {c}" for i, c in enumerate(comments)])
    full_prompt = prompt_template + comments_text + "\n\n只输出JSON数组，���要其他内容。"
    
    try:
        r = DEEPSEEK_CLIENT.chat.completions.create(
            model="deepseek-chat",
            messages=[
                {"role": "system", "content": "你是短视频评论分类专家，只输出结构化的JSON数组。"},
                {"role": "user", "content": full_prompt}
            ],
            max_tokens=2000,
            temperature=0.1,
        )
        content = r.choices[0].message.content.strip()
        if content.startswith("```"):
            content = re.sub(r'^```(?:json)?\s*', '', content)
            content = re.sub(r'\s*```$', '', content)
        results = json.loads(content)
        if isinstance(results, list):
            return {item.get("index", 0): item.get("category", "other") for item in results}
        return {}
    except Exception as e:
        print(f"[LLM fallback] Comment classification failed: {e}")
        return {}


# ============================================================
# API: Batch analyze — 批量检查缓存状态
# ============================================================

def _wrap_classify(title, lang, en_config, es_config):
    """分类器包装：支持 LLM 兜底"""
    config = es_config if lang == "es" else en_config
    normalized = re.sub(r"\s+", " ", (title or "").strip().lower())
    override = (config or {}).get("calibration_overrides", {}).get(normalized)
    if override:
        return ("", "人工校准", override.get("category", "其他"),
                "Calibration override: " + override.get("reason", ""))
    if lang == "es":
        return classify_title_es(title, es_config)
    else:
        return classify_title_en(title, en_config)


@app.route("/api/calibrate/title", methods=["POST"])
def api_calibrate_title():
    """保存视频选题校准；后续分析由统一分类包装器优先应用。"""
    try:
        data = request.get_json() or {}
        title = data.get("title", "").strip()
        category = data.get("category", "").strip()
        reason = data.get("reason", "").strip()
        language = data.get("language", "en")
        if not title or not category or not reason:
            return jsonify({"error": "标题、正确分类和校准理由均不能为空"}), 400
        if category not in CAT_COLORS:
            return jsonify({"error": f"无效分类：{category}"}), 400

        filename = "spanish_category_config.json" if language == "es" else "english_category_config.json"
        path = os.path.join(BASE_DIR, "configs", filename)
        with open(path, "r", encoding="utf-8") as f:
            config = json.load(f)

        model_rule = ""
        if DEEPSEEK_CLIENT:
            try:
                prompt = (
                    f"短视频标题应校准为“{category}”。请根据标题和人工理由，输出一句可复用分类规则。"
                    f"\n标题：{title}\n人工理由：{reason}\n只输出规则。"
                )
                model_rule = DEEPSEEK_CLIENT.chat.completions.create(
                    model="deepseek-chat",
                    messages=[{"role": "user", "content": prompt}],
                    max_tokens=180,
                    temperature=0.2,
                ).choices[0].message.content.strip()
            except Exception as e:
                model_rule = f"模型分析暂不可用：{str(e)[:80]}"

        key = re.sub(r"\s+", " ", title.lower()).strip()
        config.setdefault("calibration_overrides", {})[key] = {
            "category": category, "reason": reason, "model_rule": model_rule,
            "updated": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        }
        with open(path, "w", encoding="utf-8") as f:
            json.dump(config, f, ensure_ascii=False, indent=2)
        return jsonify({"success": True, "category": category, "model_rule": model_rule})
    except Exception as e:
        return jsonify({"error": f"选题校准失败：{str(e)}"}), 500


@app.route("/api/batch/analyze", methods=["POST"])
def api_batch_analyze():
    """检查多个达人的缓存状态"""
    try:
        data = request.get_json()
        inputs = data.get("inputs", [])
        platform = data.get("platform", "Instagram")
        days = int(data.get("days", 30))
        language = data.get("language", "en")

        cached = []
        need_apify = []

        for inp in inputs:
            username = extract_username(inp, platform)
            if not username:
                continue
            need_refresh, cached_videos, msg = needs_apify_refresh(username, platform, days)
            if need_refresh:
                need_apify.append({"username": username, "cached_count": len(cached_videos), "message": msg})
            else:
                cached.append({"username": username, "count": len(cached_videos), "message": msg})

        return jsonify({
            "total": len(inputs),
            "cached": cached,
            "need_apify": need_apify,
        })
    except Exception as e:
        import traceback
        return jsonify({"error": f"批量检查失败: {str(e)}", "trace": traceback.format_exc()}), 500


@app.route("/api/batch/fetch", methods=["POST"])
def api_batch_fetch():
    """批量 Apify 抓取 + 分类（同步，前端逐个轮询）"""
    try:
        data = request.get_json()
        usernames = data.get("usernames", [])
        platform = data.get("platform", "Instagram")
        days = int(data.get("days", 30))
        language = data.get("language", "en")

        en_config = load_en_config() if language != "es" else None
        es_config = load_es_config() if language == "es" else None

        results = []
        total_done = 0
        total_error = 0

        for username in usernames:
            try:
                cached_videos = load_cached_videos(username, platform)
                
                if platform.lower() == "instagram":
                    new_videos = fetch_instagram_videos(username, APIFY_API_KEY, days)
                else:
                    new_videos = fetch_tiktok_videos(username, APIFY_API_KEY, days)
                
                merged = merge_videos(cached_videos, new_videos)
                save_cached_videos(username, platform, merged)
                window_videos = filter_videos_by_days(merged, days)
                
                # Classify + LLM fallback
                classified = []
                unclassified_titles = []
                unclassified_indices = []
                
                for i, v in enumerate(window_videos):
                    title = v.get("标题", v.get("title", ""))
                    brand, keywords, category, basis = _wrap_classify(title, language, en_config, es_config)
                    if category == "其他" and basis == "No rule matched":
                        unclassified_titles.append(title)
                        unclassified_indices.append(i)
                    
                    classified.append({
                        "发布日期": v.get("发布日期", v.get("date", "")),
                        "达人ID": username,
                        "平台": platform,
                        "标题": title,
                        "视频链接": v.get("视频链接", v.get("url", "")),
                        "评论数": int(v.get("评论数", v.get("comments_count", 0)) or 0),
                        "分类": category,
                        "命中品牌": brand or "",
                        "命中关键词": keywords or "",
                        "分类依据": basis,
                    })
                
                # LLM 兜底：对未命中的标题做批量分类
                if en_config and unclassified_titles:
                    llm_results = _llm_classify_titles(unclassified_titles, en_config)
                    for local_idx, cat in llm_results.items():
                        if local_idx < len(unclassified_indices):
                            result_idx = unclassified_indices[local_idx]
                            classified[result_idx]["分类"] = cat
                            classified[result_idx]["分类依据"] = "LLM fallback"
                
                results.append({"username": username, "videos": classified, "error": None})
                total_done += 1
            except Exception as e:
                results.append({"username": username, "videos": [], "error": str(e)})
                total_error += 1

        return jsonify({"results": results, "total_done": total_done, "total_error": total_error})
    except Exception as e:
        import traceback
        return jsonify({"error": f"批量抓取失败: {str(e)}", "trace": traceback.format_exc()}), 500


# ============================================================
# API: Apify 异步抓取 + 状态轮询
# ============================================================

def _apify_fetch_task(task_id, username, platform, days, language):
    """异步 Apify 抓取任务"""
    try:
        _async_tasks[task_id]["status"] = "running"
        cached = load_cached_videos(username, platform)
        
        if platform.lower() == "instagram":
            new_videos = fetch_instagram_videos(username, APIFY_API_KEY, days)
        else:
            new_videos = fetch_tiktok_videos(username, APIFY_API_KEY, days)
        
        merged = merge_videos(cached, new_videos)
        save_cached_videos(username, platform, merged)
        window_videos = filter_videos_by_days(merged, days)
        
        en_config = load_en_config() if language != "es" else None
        es_config = load_es_config() if language == "es" else None
        
        classified = []
        unclassified_titles = []
        unclassified_indices = []
        
        for i, v in enumerate(window_videos):
            title = v.get("标题", v.get("title", ""))
            brand, keywords, category, basis = _wrap_classify(title, language, en_config, es_config)
            if category == "其他" and basis == "No rule matched":
                unclassified_titles.append(title)
                unclassified_indices.append(i)
            
            classified.append({
                "发布日期": v.get("发布日期", v.get("date", "")),
                "达人ID": username,
                "平台": platform,
                "标题": title,
                "视频链接": v.get("视频链接", v.get("url", "")),
                "评论数": int(v.get("评论数", v.get("comments_count", 0)) or 0),
                "分类": category,
                "命中品牌": brand or "",
                "命中关键词": keywords or "",
                "分类依据": basis,
            })
        
        # LLM 兜底
        if en_config and unclassified_titles:
            llm_results = _llm_classify_titles(unclassified_titles, en_config)
            for local_idx, cat in llm_results.items():
                if local_idx < len(unclassified_indices):
                    result_idx = unclassified_indices[local_idx]
                    classified[result_idx]["分类"] = cat
                    classified[result_idx]["分类依据"] = "LLM fallback"
        
        _async_tasks[task_id] = {
            "status": "done",
            "result": {
                "username": username,
                "platform": platform,
                "videos": classified,
                "total": len(classified),
                "message": f"Apify 抓取完成：{len(classified)} 条视频",
            }
        }
    except Exception as e:
        _async_tasks[task_id] = {"status": "error", "error": str(e)}


@app.route("/api/apify/fetch", methods=["POST"])
def api_apify_fetch():
    """启动 Apify 异步抓取任务"""
    try:
        data = request.get_json()
        username = extract_username(data.get("input", data.get("username", "")), data.get("platform", ""))
        platform = data.get("platform", "Instagram")
        days = int(data.get("days", 30))
        language = data.get("language", "en")
        
        if not username:
            return jsonify({"error": "无法解析达人ID"}), 400
        
        task_id = str(uuid.uuid4())[:8]
        _async_tasks[task_id] = {"status": "started"}
        
        threading.Thread(target=_apify_fetch_task, args=(task_id, username, platform, days, language), daemon=True).start()
        
        return jsonify({"task_id": task_id, "status": "started", "username": username})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/apify/status/<task_id>")
def api_apify_status(task_id):
    """查询 Apify 抓取任务状态"""
    task = _async_tasks.get(task_id)
    if not task:
        return jsonify({"status": "not_found", "error": "任务不存在或已过期"})
    return jsonify(task)


# ============================================================
# API: Single video analysis (Panel 2)
# ============================================================

def _video_analyze_task(task_id, url, detected_platform, language, comment_count):
    """异步单条视频分析任务"""
    try:
        _async_tasks[task_id]["status"] = "running"
        
        # Fetch metadata
        metadata = _fetch_video_metadata_sc(url, detected_platform)
        
        # Fetch transcript
        transcript = _fetch_transcript_sc(url, detected_platform)
        
        # Fetch comments
        comments_raw = _fetch_comments_single_video_sc(
            url, detected_platform, target_valid=None if comment_count == "all" else int(comment_count)
        )
        
        # Classify comments
        comment_config = load_comment_config()
        classified_comments = []
        unclassified_texts = []
        unclassified_indices = []
        
        for i, c in enumerate(comments_raw):
            text = c.get("text", "")
            cat_key, name_en, name_zh, signals = classify_single_comment(text, comment_config)
            classified_comments.append({
                "text": text[:200],
                "likes": c.get("likes", 0),
                "username": c.get("username", ""),
                "category": name_zh,
                "category_key": cat_key,
                "matched_signals": signals,
            })
            if cat_key == "other":
                unclassified_texts.append(text[:200])
                unclassified_indices.append(len(classified_comments) - 1)
        
        # LLM 兜底：评论分类
        if comment_config.get("llm_fallback", {}).get("enabled") and unclassified_texts:
            llm_results = _llm_classify_comments(unclassified_texts, comment_config)
            for idx, cat_key in llm_results.items():
                if idx < len(classified_comments):
                    cat_info = COMMENT_CAT_NAMES.get(cat_key, {})
                    classified_comments[idx]["category_key"] = cat_key
                    classified_comments[idx]["category"] = COMMENT_CAT_NAMES.get(cat_key, "其他")
        
        # DeepSeek content analysis
        comments_sample = [{"text": c.get("text", "")} for c in comments_raw[:15]]
        analysis = _analyze_video_with_deepseek(transcript, comments_sample, language)
        transcript_zh = _translate_full_text_to_chinese(transcript)
        
        _async_tasks[task_id] = {
            "status": "done",
            "result": {
                "metadata": metadata,
                "transcript": transcript or "",
                "transcript_zh": transcript_zh,
                "analysis": analysis,
                "comments": classified_comments,
                "total_comments": len(comments_raw),
                "requested_comment_count": comment_count,
            }
        }
    except Exception as e:
        _async_tasks[task_id] = {"status": "error", "error": str(e)}


@app.route("/api/video/analyze", methods=["POST"])
def api_video_analyze():
    """启动单条视频分析"""
    try:
        data = request.get_json()
        url = data.get("url", "").strip()
        platform = data.get("platform", "")
        language = data.get("language", "en")
        raw_comment_count = data.get("comment_count", 50)
        comment_count = "all" if str(raw_comment_count).lower() == "all" else max(0, min(int(raw_comment_count), 500))
        
        if not url:
            return jsonify({"error": "请输入视频链接"}), 400
        
        detected_platform = platform or _detect_platform(url)
        if not detected_platform:
            return jsonify({"error": "无法识别平台，请手动选择"}), 400
        if not _is_single_video_url(url, detected_platform):
            if detected_platform == "Instagram":
                return jsonify({
                    "error": "这不是单条 Instagram 视频链接。请打开具体视频后，粘贴形如 https://www.instagram.com/reel/视频短码/ 的链接；达人主页或 /reels/ 列表页不能用于单视频分析。"
                }), 400
            return jsonify({
                "error": "这不是单条 TikTok 视频链接。请粘贴包含 /@账号/video/视频ID 的具体视频链接。"
            }), 400
        
        task_id = str(uuid.uuid4())[:8]
        _async_tasks[task_id] = {"status": "started"}
        
        threading.Thread(
            target=_video_analyze_task,
            args=(task_id, url, detected_platform, language, comment_count),
            daemon=True,
        ).start()
        
        return jsonify({"task_id": task_id, "status": "started"})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/video/status/<task_id>")
def api_video_status(task_id):
    """查询视频分析任务状态"""
    task = _async_tasks.get(task_id)
    if not task:
        return jsonify({"status": "not_found", "error": "任务不存在或已过期"})
    return jsonify(task)


# ============================================================
# API: Paste Views (Panel 3)
# ============================================================

def _split_pasted_column(raw):
    """按 Excel 粘贴行原样拆分，保留中间空行。"""
    if raw is None or raw == "":
        return []
    return str(raw).replace("\r\n", "\n").replace("\r", "\n").split("\n")


def _parse_view_value(value):
    text = str(value or "").strip().replace(",", "")
    if not text:
        return None
    try:
        return max(0, int(float(text)))
    except (TypeError, ValueError):
        return None


def _max_view(live_value, historical_value):
    values = [v for v in (live_value, historical_value) if isinstance(v, (int, float))]
    return int(max(values)) if values else None


def _paste_views_task(task_id, tk_urls, tk_views, ig_urls, ig_views):
    """异步批量抓取 View 数据"""
    try:
        _async_tasks[task_id]["status"] = "running"
        tk_lines = _split_pasted_column(tk_urls)
        tk_history = _split_pasted_column(tk_views)
        ig_lines = _split_pasted_column(ig_urls)
        ig_history = _split_pasted_column(ig_views)
        
        total_rows = max(len(tk_lines), len(tk_history), len(ig_lines), len(ig_history))
        rows = []
        
        for i in range(total_rows):
            row = {
                "row": i + 1,
                "tk_live_view": None, "tk_history_view": _parse_view_value(tk_history[i]) if i < len(tk_history) else None,
                "ig_live_view": None, "ig_history_view": _parse_view_value(ig_history[i]) if i < len(ig_history) else None,
                "tk_view": None, "ig_view": None, "tk_error": None, "ig_error": None,
            }
            
            if i < len(tk_lines) and tk_lines[i]:
                try:
                    meta = _fetch_video_metadata_sc(tk_lines[i], "TikTok")
                    row["tk_live_view"] = _parse_view_value(meta.get("views", 0))
                except Exception as e:
                    row["tk_error"] = str(e)[:100]
            
            if i < len(ig_lines) and ig_lines[i]:
                try:
                    meta = _fetch_video_metadata_sc(ig_lines[i], "Instagram")
                    row["ig_live_view"] = _parse_view_value(meta.get("views", 0))
                except Exception as e:
                    row["ig_error"] = str(e)[:100]

            row["tk_view"] = _max_view(row["tk_live_view"], row["tk_history_view"])
            row["ig_view"] = _max_view(row["ig_live_view"], row["ig_history_view"])
            
            rows.append(row)
        
        _async_tasks[task_id] = {"status": "done", "result": {"rows": rows, "total": total_rows}}
    except Exception as e:
        _async_tasks[task_id] = {"status": "error", "error": str(e)}


@app.route("/api/videos/views/paste", methods=["POST"])
def api_paste_views():
    """批量 paste 链接，抓取 View"""
    try:
        data = request.get_json()
        tk_urls = data.get("tk_urls", "")
        tk_views = data.get("tk_views", "")
        ig_urls = data.get("ig_urls", "")
        ig_views = data.get("ig_views", "")
        
        total = max(
            len(_split_pasted_column(tk_urls)), len(_split_pasted_column(tk_views)),
            len(_split_pasted_column(ig_urls)), len(_split_pasted_column(ig_views)),
        )
        
        if total == 0:
            return jsonify({"error": "请至少粘贴一条链接"}), 400
        
        task_id = str(uuid.uuid4())[:8]
        _async_tasks[task_id] = {"status": "started"}
        
        threading.Thread(
            target=_paste_views_task,
            args=(task_id, tk_urls, tk_views, ig_urls, ig_views),
            daemon=True,
        ).start()
        
        return jsonify({"task_id": task_id, "status": "started", "total": total})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/videos/views/status/<task_id>")
def api_paste_views_status(task_id):
    """查询 View 抓取状态"""
    task = _async_tasks.get(task_id)
    if not task:
        return jsonify({"status": "not_found", "error": "任务不存在或已过期"})
    return jsonify(task)


# ============================================================
# API: Upload videos
# ============================================================

@app.route("/api/upload/videos", methods=["POST"])
def api_upload_videos():
    """上传视频 JSON 数据"""
    try:
        data = request.get_json()
        username = data.get("username", "")
        platform = data.get("platform", "Instagram")
        videos = data.get("videos", [])
        
        if not username or not videos:
            return jsonify({"error": "缺少必要参数"}), 400
        
        # Normalize fields
        normalized = []
        for v in videos:
            normalized.append({
                "发布日期": v.get("date", v.get("发布日期", "")),
                "标题": v.get("title", v.get("标题", "")),
                "视频链接": v.get("url", v.get("视频链接", "")),
                "评论数": int(v.get("comments_count", v.get("评论数", 0)) or 0),
            })
        
        save_cached_videos(username, platform, normalized)
        
        return jsonify({
            "message": f"已保存 {len(normalized)} 条视频",
            "videos": normalized,
            "username": username,
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ============================================================
# API: Excel Download
# ============================================================

@app.route("/api/download/excel", methods=["POST"])
def api_download_excel():
    """生成并下载 Excel 报告"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from io import BytesIO
        
        data = request.get_json()
        videos = data.get("videos", [])
        comments_data = data.get("comments", {})
        username = data.get("username", "unknown")
        platform = data.get("platform", "")
        
        wb = openpyxl.Workbook()
        
        # Sheet 1: 视频选题分析
        ws1 = wb.active
        ws1.title = "视频选题分析"
        headers = ["发布日期", "达人ID", "平台", "标题", "视频链接", "评论数", "分类", "命中品牌", "命中关键词", "分类依据"]
        for col, h in enumerate(headers, 1):
            cell = ws1.cell(row=1, column=col, value=h)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        for row_idx, v in enumerate(videos, 2):
            ws1.cell(row=row_idx, column=1, value=v.get("发布日期", ""))
            ws1.cell(row=row_idx, column=2, value=v.get("达人ID", username))
            ws1.cell(row=row_idx, column=3, value=v.get("平台", platform))
            ws1.cell(row=row_idx, column=4, value=v.get("标题", ""))
            ws1.cell(row=row_idx, column=5, value=v.get("视频链接", ""))
            ws1.cell(row=row_idx, column=6, value=v.get("评论数", 0))
            ws1.cell(row=row_idx, column=7, value=v.get("分类", ""))
            ws1.cell(row=row_idx, column=8, value=v.get("命中品牌", ""))
            ws1.cell(row=row_idx, column=9, value=v.get("命中关键词", ""))
            ws1.cell(row=row_idx, column=10, value=v.get("分类依据", ""))
        
        ws1.column_dimensions['D'].width = 40
        ws1.column_dimensions['E'].width = 35
        
        # Sheet 2: 评论分析（Part 1 + Part 2）
        ws2 = wb.create_sheet("评论分析")
        row = 1
        
        for part_key, part_title in [("part1_top3_hot", "Part 1: TOP-3 热门视频评论"), ("part2_top3_sponsored", "Part 2: TOP-3 赞助/种草视频评论")]:
            part = comments_data.get(part_key, {})
            if not part:
                continue
            
            cell = ws2.cell(row=row, column=1, value=part_title)
            cell.font = Font(bold=True, size=14)
            row += 1
            
            cell = ws2.cell(row=row, column=1, value=f"购买意向评分: {part.get('purchase_intent_score', 'N/A')} ({part.get('purchase_intent_label', '')})")
            cell.font = Font(bold=True, color="4472C4")
            row += 2
            
            for v in part.get("videos", []):
                ws2.cell(row=row, column=1, value=f"视频: {v.get('title', '')[:80]}")
                ws2.cell(row=row, column=1).font = Font(bold=True)
                row += 1
                
                # Comment headers
                for col, h in enumerate(["评论内容", "点赞", "用户名", "分类"], 1):
                    cell = ws2.cell(row=row, column=col, value=h)
                    cell.font = Font(bold=True)
                row += 1
                
                for c in v.get("classified_comments", []):
                    ws2.cell(row=row, column=1, value=c.get("text", "")[:200])
                    ws2.cell(row=row, column=2, value=c.get("likes", 0))
                    ws2.cell(row=row, column=3, value=c.get("username", ""))
                    ws2.cell(row=row, column=4, value=c.get("category", ""))
                    row += 1
                row += 1
        
        ws2.column_dimensions['A'].width = 50
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        from flask import send_file
        filename = f"达人分析_{username}_{platform}_{datetime.now().strftime('%Y%m%d')}.xlsx"
        return send_file(
            output,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=filename,
        )
    except Exception as e:
        import traceback
        return jsonify({"error": f"Excel 生成失败: {str(e)}", "trace": traceback.format_exc()}), 500


# ============================================================
# API: Config Management
# ============================================================

@app.route("/api/config/keys")
def api_config_keys():
    """返回 API Key 设置状态（不暴露完整 Key）"""
    def _mask(k):
        if not k:
            return "未设置"
        if len(k) <= 8:
            return k[:2] + "***"
        return k[:4] + "***" + k[-4:]
    
    return jsonify({
        "scrapecreators": {
            "set": bool(SC_API_KEYS and SC_API_KEYS[0]),
            "masked": _mask(SC_API_KEYS[0]) if SC_API_KEYS else "",
            "count": len(SC_API_KEYS),
        },
        "apify": {
            "set": bool(APIFY_API_KEY),
            "masked": _mask(APIFY_API_KEY),
        },
        "deepseek": {
            "set": bool(DEEPSEEK_API_KEY),
            "masked": _mask(DEEPSEEK_API_KEY),
        }
    })


@app.route("/api/config/credits")
def api_config_credits():
    """查询各 API 额度"""
    try:
        sc = _check_sc_credits()
    except Exception as e:
        sc = {"keys": [], "active": 0, "error": str(e)[:100]}
    try:
        ap = _check_apify_credits()
    except Exception as e:
        ap = {"remaining": 0, "error": str(e)[:100]}
    try:
        ds = _check_deepseek_status()
    except Exception as e:
        ds = {"status": "unavailable", "error": str(e)[:100]}
    try:
        return jsonify({"scrapecreators": sc, "apify": ap, "deepseek": ds})
    except Exception:
        return jsonify({"scrapecreators": {"keys": [], "active": 0}, "apify": None, "deepseek": {"status": "unavailable"}})


@app.route("/api/config/scrapecreators-key", methods=["POST"])
def api_set_sc_key():
    try:
        data = request.get_json()
        key = data.get("key", "").strip()
        if not key:
            return jsonify({"error": "Key 不能为空"}), 400
        # Update api_keys.py
        _api_keys.SCRAPECREATORS_API_KEY = key
        with open(os.path.join(BASE_DIR, "configs", "api_keys.py"), "w", encoding="utf-8") as f:
            f.write(f'# API Keys for influencer-hub\n')
            f.write(f'SCRAPECREATORS_API_KEY = """{key}"""\n')
            f.write(f'DEEPSEEK_API_KEY = """{DEEPSEEK_API_KEY}"""\n')
            f.write(f'APIFY_API_KEY = """{APIFY_API_KEY}"""\n')
        reload_api_keys()
        return jsonify({"status": "ok"})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/config/apify-key", methods=["POST"])
def api_set_apify_key():
    try:
        data = request.get_json()
        key = data.get("key", "").strip()
        if not key:
            return jsonify({"error": "Key 不能为空"}), 400
        _api_keys.APIFY_API_KEY = key
        with open(os.path.join(BASE_DIR, "configs", "api_keys.py"), "w", encoding="utf-8") as f:
            f.write(f'# API Keys for influencer-hub\n')
            f.write(f'SCRAPECREATORS_API_KEY = """{SC_API_KEY}"""\n')
            if len(SC_API_KEYS) > 1:
                f.write(f'# 多 Key 支持：\n')
                f.write(f'SCRAPECREATORS_API_KEY = """{",".join(SC_API_KEYS)}"""\n')
            else:
                f.write(f'SCRAPECREATORS_API_KEY = """{SC_API_KEY}"""\n')
            f.write(f'DEEPSEEK_API_KEY = """{DEEPSEEK_API_KEY}"""\n')
            f.write(f'APIFY_API_KEY = """{key}"""\n')
        reload_api_keys()
        return jsonify({"status": "ok"})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/config/deepseek-key", methods=["POST"])
def api_set_deepseek_key():
    try:
        data = request.get_json()
        key = data.get("key", "").strip()
        if not key:
            return jsonify({"error": "Key 不能为空"}), 400
        _api_keys.DEEPSEEK_API_KEY = key
        with open(os.path.join(BASE_DIR, "configs", "api_keys.py"), "w", encoding="utf-8") as f:
            f.write(f'# API Keys for influencer-hub\n')
            if len(SC_API_KEYS) > 1:
                f.write(f'SCRAPECREATORS_API_KEY = """{",".join(SC_API_KEYS)}"""\n')
            else:
                f.write(f'SCRAPECREATORS_API_KEY = """{SC_API_KEY}"""\n')
            f.write(f'DEEPSEEK_API_KEY = """{key}"""\n')
            f.write(f'APIFY_API_KEY = """{APIFY_API_KEY}"""\n')
        reload_api_keys()
        return jsonify({"status": "ok"})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ============================================================
# API: Panel 4 — 批量评论洞察（视频链接分组对比）
# ============================================================

@app.route("/api/batch/comments", methods=["POST"])
def api_batch_comments():
    """
    按垂类达人 / 生活种草达人两组视频链接，抓取评论并分类对比。
    Input: {
        vertical_links: [...],
        lifestyle_links: [...],
        platform: "Instagram|TikTok",
        comment_count: 20
    }
    Output: {
        vertical: {total_videos, total_comments, distribution, top_comments},
        lifestyle: {total_videos, total_comments, distribution, top_comments},
        insights: {text}  // LLM 生成的对比洞察
    }
    """
    try:
        data = request.get_json()
        vertical_links = data.get("vertical_links", [])
        lifestyle_links = data.get("lifestyle_links", [])
        platform = data.get("platform", "Instagram")
        comment_count = int(data.get("comment_count", 20))

        if not vertical_links and not lifestyle_links:
            return jsonify({"error": "请至少输入一条视频链接"}), 400

        comment_config = load_comment_config()

        # --- Analyze one group of video links ---
        def analyze_group(links, group_name):
            """对一组视频链接并行抓取评论、分类、汇总"""
            all_classified = []
            successful_videos = 0
            video_errors = []

            def analyze_single_video(url):
                nonlocal successful_videos
                nonlocal all_classified
                result = {"url": url, "classified": [], "error": None, "title": ""}
                try:
                    comments_raw = fetch_comments_for_video(url, platform)
                    valid = get_top_valid_comments(comments_raw, comment_count)
                    for c in valid:
                        cat_key, name_en, name_zh, signals = classify_single_comment(c["text"], comment_config)
                        result["classified"].append({
                            "text": c["text"][:200],
                            "likes": c["likes"],
                            "username": c.get("username", ""),
                            "category_key": cat_key,
                            "category": name_zh,
                            "matched_signals": signals,
                        })
                    successful_videos += 1
                except Exception as e:
                    result["error"] = str(e)
                return result

            # Parallel fetch per video (max 4 concurrent)
            with ThreadPoolExecutor(max_workers=min(4, len(links))) as ex:
                futures = {ex.submit(analyze_single_video, url): url for url in links if url.strip()}
                for f in futures:
                    r = f.result()
                    if r["error"]:
                        video_errors.append({"url": r["url"], "error": r["error"]})
                    all_classified.extend(r["classified"])

            # LLM 兜底：对 other 评论尝试重新分类
            if comment_config.get("llm_fallback", {}).get("enabled", False):
                unclassified = [(i, c) for i, c in enumerate(all_classified) if c["category_key"] == "other"]
                if unclassified:
                    texts = [c[1]["text"] for c in unclassified]
                    try:
                        llm_results = _llm_classify_comments(texts, comment_config)
                        for local_idx, cat_key in llm_results.items():
                            if local_idx < len(unclassified):
                                orig_idx = unclassified[local_idx][0]
                                all_classified[orig_idx]["category_key"] = cat_key
                                all_classified[orig_idx]["category"] = COMMENT_CAT_NAMES.get(cat_key, "其他")
                    except Exception:
                        pass

            # Distribution
            dist = {}
            for k in ["content_engagement", "purchase_intent", "product_interaction", "other"]:
                dist[k] = {"count": 0, "pct": 0, "name_zh": COMMENT_CAT_NAMES.get(k, k)}

            for c in all_classified:
                ck = c["category_key"]
                if ck in dist:
                    dist[ck]["count"] += 1

            total = len(all_classified)
            for k in dist:
                dist[k]["pct"] = round(dist[k]["count"] / total * 100, 1) if total > 0 else 0

            # Purchase intent score
            score, label = _calc_purchase_intent(all_classified, comment_config)

            # Top purchase intent comments
            pi_comments = [c for c in all_classified if c["category_key"] == "purchase_intent"]
            pi_comments.sort(key=lambda c: c.get("likes", 0), reverse=True)

            return {
                "group": group_name,
                "total_videos": successful_videos,
                "total_links": len(links),
                "total_comments": total,
                "distribution": dist,
                "purchase_intent_score": score,
                "purchase_intent_label": label,
                "top_comments": pi_comments[:10],
                "all_comments": all_classified,
                "errors": video_errors[:5],
            }

        # Analyze both groups in parallel
        vertical_result = None
        lifestyle_result = None
        with ThreadPoolExecutor(max_workers=2) as ex:
            fv = ex.submit(analyze_group, vertical_links, "垂类达人") if vertical_links else None
            fl = ex.submit(analyze_group, lifestyle_links, "生活种草达人") if lifestyle_links else None
            if fv:
                vertical_result = fv.result()
            if fl:
                lifestyle_result = fl.result()

        # LLM 生成对比洞察
        insights = None
        if vertical_result and lifestyle_result:
            try:
                insights = _llm_generate_comparison_insights(vertical_result, lifestyle_result)
            except Exception:
                insights = {"text": "LLM 洞察生成失败，请查看下方统计数据。"}

        return jsonify({
            "vertical": vertical_result,
            "lifestyle": lifestyle_result,
            "insights": insights,
            "platform": platform,
        })
    except Exception as e:
        import traceback
        return jsonify({"error": f"批量评论分析失败: {str(e)}", "trace": traceback.format_exc()}), 500


# ============ 评论分类在线校准 ============

COMMENT_CAT_ORDER = ["content_engagement", "purchase_intent", "product_interaction", "other"]
COMMENT_CAT_NAMES = {
    "content_engagement": "内容互动",
    "purchase_intent": "购买意向",
    "product_interaction": "产品互动",
    "other": "其他",
}


def _cal_norm(text):
    return " " + (text or "").lower().strip() + " "


def _cal_hit(text_norm, signals):
    for s in signals:
        if s.lower() in text_norm:
            return s
    return None


def _cal_extract_keyword(text, target_cat, config):
    """从评论中提取最特异的短语用于加入目标分类信号"""
    words = text.lower().strip().split()
    if not words:
        return None

    # 收集所有已有信号（避免重复）
    all_signals = set()
    for ck, cc in config.get("categories", {}).items():
        for s in cc.get("signals", []):
            all_signals.add(s.lower())

    # 尝试 3-gram → 2-gram，选最长的全新短语
    for n in [3, 2]:
        for i in range(len(words) - n + 1):
            phrase = " ".join(words[i:i+n])
            if len(phrase) >= 5 and phrase not in all_signals:
                return phrase
    return None


def _cal_extract_keyword_force(text):
    """兜底：强制取评论最前面3个词"""
    words = text.lower().strip().split()
    if len(words) >= 3:
        return " ".join(words[:3])
    return text.lower().strip()


def _cal_diagnose(text, old_cat, new_cat, config):
    """分析错分原因，返回诊断文本列表"""
    t = _cal_norm(text)
    cats = config.get("categories", {})
    order = config.get("classification_order", COMMENT_CAT_ORDER)
    lines = []

    # 1. 逐类解释为什么到了 old_cat
    old_idx = order.index(old_cat) if old_cat in order else len(order)
    for ck in order:
        cat = cats.get(ck, {})
        if ck == new_cat or (ck == old_cat and ck != "other"):
            break
        matched_sig = _cal_hit(t, cat.get("signals", []))
        if matched_sig:
            lines.append(f"命中了「{cat.get('name_zh', ck)}」的信号「{matched_sig}」→ 被优先归为该类")
        elif ck == "other":
            if old_cat == "other":
                lines.append("未命中任何分类信号 → 归入「其他」")

    # 2. 检查目标分类是否有匹配信号（但被 exclude 扼杀）
    target = cats.get(new_cat, {})
    matched_target = _cal_hit(t, target.get("signals", []))
    matched_exclude = _cal_hit(t, target.get("exclude_signals", []))

    if matched_target:
        if lines:
            lines.append(f"目标分类「{target.get('name_zh', new_cat)}」有匹配信号「{matched_target}」，但被优先级更高的类抢走")
        else:
            lines.append(f"目标分类「{target.get('name_zh', new_cat)}」有匹配信号「{matched_target}」（当前分类规则认为已匹配，但实际被归入了 {old_cat}）")
    elif matched_exclude:
        lines.append(f"目标分类「{target.get('name_zh', new_cat)}」的排除信号「{matched_exclude}」错误排除了该评论")
    else:
        lines.append(f"目标分类「{target.get('name_zh', new_cat)}」无匹配信号，需新增关键词")

    return lines, matched_exclude


def _cal_apply_fix(text, new_cat, matched_exclude, config):
    """应用修复：移除误排除、新增关键词信号"""
    t = _cal_norm(text)
    cats = config.get("categories", {})
    target = cats.get(new_cat, {})
    fix_items = []

    # Fix 1: 移除误排除的 exclude_signal
    if matched_exclude:
        excludes = target.get("exclude_signals", [])
        if matched_exclude in excludes:
            target["exclude_signals"] = [s for s in excludes if s != matched_exclude]
            fix_items.append(f"移除排除信号「{matched_exclude}」")

    # Fix 2: 新增关键词（仅当目标分类没有任何匹配信号时）
    matched_target = _cal_hit(t, target.get("signals", []))
    keyword = None
    if not matched_target:
        keyword = _cal_extract_keyword(text, new_cat, config)
        if not keyword:
            keyword = _cal_extract_keyword_force(text)
        existing = [s.lower() for s in target.get("signals", [])]
        if keyword and keyword.lower() not in existing:
            target["signals"].append(keyword.lower())
            fix_items.append(f"新增信号「{keyword}」")

    return fix_items, keyword


@app.route("/api/batch/calibrate_comment", methods=["POST"])
def api_calibrate_comment():
    """
    校准单条评论分类。自动分析错分原因并更新配置文件。
    Input: {text: str, old_category: str, new_category: str}
    """
    try:
        from engines.comment_analyzer import load_config, classify_single_comment
        import engines.comment_analyzer as ca_mod
        import copy

        data = request.get_json()
        text = data.get("text", "").strip()
        old_cat = data.get("old_category", "other")
        new_cat = data.get("new_category", "purchase_intent")
        reason = data.get("reason", "").strip()

        if not text:
            return jsonify({"error": "评论原文不能为空"}), 400
        if new_cat not in COMMENT_CAT_NAMES:
            return jsonify({"error": f"无效的目标分类: {new_cat}"}), 400
        if not reason:
            return jsonify({"error": "请填写校准理由"}), 400

        config = load_config()

        # 加载前保存备份
        config_backup = copy.deepcopy(config)

        # Step 1: 诊断
        diagnosis_lines, matched_exclude = _cal_diagnose(text, old_cat, new_cat, config)
        diagnosis_lines.insert(0, f"人工校准理由：{reason}")

        # Step 2: 应用修复
        fix_items, keyword = _cal_apply_fix(text, new_cat, matched_exclude, config)

        model_rule = ""
        if DEEPSEEK_CLIENT:
            try:
                prompt = (
                    f"评论应校准为“{COMMENT_CAT_NAMES[new_cat]}”。请根据评论和人工理由，"
                    f"输出一句可复用分类规则。\n评论：{text}\n人工理由：{reason}\n只输出规则。"
                )
                model_rule = DEEPSEEK_CLIENT.chat.completions.create(
                    model="deepseek-chat",
                    messages=[{"role": "user", "content": prompt}],
                    max_tokens=180,
                    temperature=0.2,
                ).choices[0].message.content.strip()
            except Exception as e:
                model_rule = f"模型分析暂不可用：{str(e)[:80]}"

        normalized = re.sub(r"\s+", " ", text.lower()).strip()
        config.setdefault("calibration_overrides", {})[normalized] = {
            "category": new_cat,
            "reason": reason,
            "model_rule": model_rule,
            "updated": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        }
        fix_items.append("保存人工校准覆盖规则")

        # Step 3: 保存配置
        config["updated"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        config_file = os.path.join(
            os.path.dirname(os.path.abspath(__file__)),
            "configs", "comment_classification_config.json"
        )
        with open(config_file, 'w', encoding='utf-8') as f:
            json.dump(config, f, indent=2, ensure_ascii=False)

        # Step 4: 清除模块缓存
        ca_mod._config_cache = config
        ca_mod._compiled_patterns_cache = {}

        # Step 5: 验证
        cat_key2, _, name_zh2, _ = classify_single_comment(text, config)
        verified = cat_key2 == new_cat

        return jsonify({
            "success": verified,
            "verified": verified,
            "reclassified_as": name_zh2,
            "diagnosis": "\n".join(diagnosis_lines),
            "fix": fix_items,
            "keyword": keyword,
            "reason": reason,
            "model_rule": model_rule,
        })

    except Exception as e:
        import traceback
        return jsonify({"error": f"校准失败: {str(e)}", "trace": traceback.format_exc()}), 500


@app.route("/api/batch/calibrate_batch", methods=["POST"])
def api_calibrate_batch():
    """批量校准多条评论"""
    try:
        data = request.get_json()
        corrections = data.get("corrections", [])
        if not corrections:
            return jsonify({"error": "校准列表为空"}), 400

        results = []
        for idx, corr in enumerate(corrections):
            text = corr.get("text", "").strip()
            old_cat = corr.get("old_category", "other")
            new_cat = corr.get("new_category", "purchase_intent")
            reason = corr.get("reason", "").strip()

            if not text:
                results.append({"index": idx, "success": False, "error": "评论原文为空"})
                continue

            from engines.comment_analyzer import load_config, classify_single_comment
            import engines.comment_analyzer as ca_mod

            config = load_config()
            diagnosis_lines, matched_exclude = _cal_diagnose(text, old_cat, new_cat, config)
            if reason:
                diagnosis_lines.insert(0, f"人工修正理由：{reason}")
            fix_items, keyword = _cal_apply_fix(text, new_cat, matched_exclude, config)

            if not fix_items:
                results.append({
                    "index": idx, "success": False,
                    "error": "无需修改配置", "text": text[:60],
                })
                continue

            config["updated"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            config_file = os.path.join(
                os.path.dirname(os.path.abspath(__file__)),
                "configs", "comment_classification_config.json"
            )
            with open(config_file, 'w', encoding='utf-8') as f:
                json.dump(config, f, indent=2, ensure_ascii=False)
            ca_mod._config_cache = config
            ca_mod._compiled_patterns_cache = {}

            cat_key2, _, name_zh2, _ = classify_single_comment(text, config)
            results.append({
                "index": idx, "success": cat_key2 == new_cat,
                "fix": fix_items, "keyword": keyword,
                "reclassified_as": name_zh2,
                "text": text[:60],
            })

        total = len(results)
        ok = sum(1 for r in results if r.get("success"))
        return jsonify({"total": total, "fixed": ok, "results": results})

    except Exception as e:
        import traceback
        return jsonify({"error": f"批量校准失败: {str(e)}", "trace": traceback.format_exc()}), 500


def _llm_generate_comparison_insights(vertical, lifestyle):
    """用 DeepSeek 生成两组评论的对比洞察"""
    if not DEEPSEEK_CLIENT:
        return {"text": "DeepSeek API Key 未配置，无法生成洞察。"}

    # Build comparison prompts
    v_pi_pct = vertical.get("distribution", {}).get("purchase_intent", {}).get("pct", 0)
    l_pi_pct = lifestyle.get("distribution", {}).get("purchase_intent", {}).get("pct", 0)
    v_dist = ", ".join(
        f"{COMMENT_CAT_NAMES.get(k,k)}:{d['count']}条({d['pct']}%)"
        for k, d in vertical.get("distribution", {}).items()
    )
    l_dist = ", ".join(
        f"{COMMENT_CAT_NAMES.get(k,k)}:{d['count']}条({d['pct']}%)"
        for k, d in lifestyle.get("distribution", {}).items()
    )
    v_top = "\n".join([f"- {c['text'][:80]}" for c in vertical.get("top_comments", [])[:5]] or ["无"])
    l_top = "\n".join([f"- {c['text'][:80]}" for c in lifestyle.get("top_comments", [])[:5]] or ["无"])

    prompt = f"""你是社交媒体评论分析师。请对比以下两组达人的评论区数据，给出中文洞察。

【垂类达人】
评论统计: {v_dist}
总评论数: {vertical.get('total_comments', 0)}
购买意向 TOP 评论:
{v_top}

【生活种草达人】
评论统计: {l_dist}
总评论数: {lifestyle.get('total_comments', 0)}
购买意向 TOP 评论:
{l_top}

请分点回复（3段，每段1-2句话）：
1. 购买意向对比（两组差异 + 原因推测）
2. 评论互动特征差异
3. 运营建议

JSON 格式输出：{{"text": "完整洞察文本"}}"""

    try:
        r = DEEPSEEK_CLIENT.chat.completions.create(
            model="deepseek-chat",
            messages=[{"role": "user", "content": prompt}],
            temperature=0.5,
            max_tokens=400,
        )
        content = r.choices[0].message.content.strip()
        if "{" in content and "}" in content:
            json_match = content[content.index("{"):content.rindex("}") + 1]
            return json.loads(json_match)
        return {"text": content}
    except Exception as e:
        return {"text": f"洞察生成异常: {str(e)}"}


# ============================================================
# App Entry
# ============================================================

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 8504))
    print(f"🚀 达人分析系统 v4.0 启动在端口 {port}")
    app.run(host="0.0.0.0", port=port, debug=True)
